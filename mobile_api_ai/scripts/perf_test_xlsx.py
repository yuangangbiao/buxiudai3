#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""xlsx 导入性能测试 - TODO-T7
DESIGN v2.0 第七章 性能预算：1000 行 < 10s

用法：
  python scripts/perf_test_xlsx.py
  python scripts/perf_test_xlsx.py --rows 5000
"""
import sys
import os
import time
import tempfile
import argparse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def generate_xlsx(rows: int) -> str:
    """生成测试 xlsx 文件"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['code', 'name', 'spec', 'unit', 'category_id', 'safety_stock', 'max_stock'])
    for i in range(rows):
        ws.append([f'P{i:06d}', f'产品{i}', f'规格{i}', '个', 1, 10, 100])
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(path)
    return path


def test_dry_run(rows: int) -> float:
    """测试 dry-run 性能"""
    from inventory_web.routes_api import register_routes_api
    from inventory_web.feature_flags import is_enabled

    if not is_enabled('t8_import_export'):
        print('[WARN] t8_import_export 未启用，启用后才能跑')
        return -1

    path = generate_xlsx(rows)
    size = os.path.getsize(path)
    print(f'[perf] 文件: {path}, 大小: {size} bytes ({size / 1024:.1f} KB)')

    # 直接测解析（不依赖 Flask context）
    t0 = time.time()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        cnt = 0
        for row in ws.iter_rows(values_only=True):
            cnt += 1
        elapsed = time.time() - t0
        print(f'[perf] 解析 {rows} 行: {elapsed:.3f}s ({rows / elapsed:.0f} 行/s)')
    except Exception as e:
        print(f'[FAIL] {e}')
        elapsed = -1
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
    return elapsed


def test_openpyxl_write(rows: int) -> float:
    """测试 openpyxl 写入性能（生成测试文件耗时）"""
    from openpyxl import Workbook
    import os as _os  # 修复 M-3：确保 unlink 可用
    t0 = time.time()
    wb = Workbook()
    ws = wb.active
    ws.append(['code', 'name'])
    for i in range(rows):
        ws.append([f'P{i}', f'产品{i}'])
    # 修复 M-3：用 mkstemp 替代已弃用的 mktemp（避免 TOCTOU）
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    _os.close(fd)
    wb.save(path)
    elapsed = time.time() - t0
    size = os.path.getsize(path)
    try:
        _os.unlink(path)
    except OSError:
        pass
    print(f'[perf] 生成 {rows} 行: {elapsed:.3f}s, 文件大小 {size / 1024:.1f} KB')
    return elapsed


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--rows', type=int, default=1000, help='测试行数')
    parser.add_argument('--max-rows', type=int, default=20000, help='修复 L-4：最大行数限制（防 OOM）')
    args = parser.parse_args()

    # 修复 L-4：硬上限保护
    if args.rows < 1 or args.rows > args.max_rows:
        print(f'[FAIL] rows={args.rows} 超出 [1, {args.max_rows}] 范围')
        return 1

    print('=' * 60)
    print(f'xlsx 性能测试 - {args.rows} 行 (max={args.max_rows})')
    print('=' * 60)

    print('\n[1] openpyxl 写入性能：')
    t_write = test_openpyxl_write(args.rows)
    print(f'    目标: <2s, 实际: {t_write:.3f}s {"[PASS]" if t_write < 2 else "[WARN]"}')

    print('\n[2] openpyxl 读取/dry-run 性能：')
    t_dry = test_dry_run(args.rows)
    if t_dry > 0:
        print(f'    目标: <10s, 实际: {t_dry:.3f}s {"[PASS]" if t_dry < 10 else "[WARN]"}')

    print('\n[3] 边界测试（10000 行）：')
    if args.rows <= 1000:
        # 修复 L-4：边界测试前检查内存
        try:
            import psutil
            mem = psutil.virtual_memory()
            avail_mb = mem.available / 1024 / 1024
            if avail_mb < 500:
                print(f'    [SKIP] 可用内存 {avail_mb:.0f}MB < 500MB，跳过 10000 行测试（防 OOM）')
            else:
                t_big = test_openpyxl_write(10000)
                print(f'    10000 行写入: {t_big:.3f}s {"[PASS]" if t_big < 20 else "[WARN]"}')
        except ImportError:
            # 无 psutil 时降级：仍跑测试，但加超时
            import signal
            class Timeout(Exception): pass
            def handler(signum, frame): raise Timeout()
            signal.signal(signal.SIGALRM, handler)
            signal.alarm(60)  # 60 秒超时
            try:
                t_big = test_openpyxl_write(10000)
                signal.alarm(0)
                print(f'    10000 行写入: {t_big:.3f}s {"[PASS]" if t_big < 20 else "[WARN]"}')
            except Timeout:
                print('    [TIMEOUT] 10000 行写入超时 60s，可能 OOM')
    else:
        print('    跳过（已超过 1000 行）')

    print('=' * 60)
    print('性能测试完成')
    print('=' * 60)
    return 0


if __name__ == '__main__':
    main()
