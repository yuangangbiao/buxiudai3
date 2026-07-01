# -*- coding: utf-8 -*-
"""库存管理 — 通用 API（已废弃，所有端点已迁移至更专业模块）

历史遗留代码已废弃（2026-06-02 安全加固审计发现）：
- PATCH/DELETE CRUD 端点已迁移到 routes_data.py
- settings 已迁移到 routes_system.py（含 admin_required）
- alert/cleanup 端点已迁移到 routes_core.py
- 保留本文件仅为防止旧前端 URL 404，所有端点返回 410 Gone

如需功能恢复，请从 routes_data.py / routes_core.py / routes_system.py 调用。
"""
import json
from flask import jsonify, render_template, request, session
from .admin_auth import admin_required, require_csrf
from .db_utils import execute, log_operation
# C-2 修复：包级 import 失败不阻断路由加载，改为 try/except 兜底
try:
    from .feature_flags import safe_require_feature as require_feature
except Exception:  # noqa: BLE001
    from functools import wraps
    def require_feature(name):  # type: ignore
        def _d(f):
            @wraps(f)
            def _w(*a, **kw):
                return f(*a, **kw)
            return _w
        return _d
    import logging
    logging.getLogger(__name__).exception('[C-2] feature_flags 加载失败，已降级为 no-op')


def register_routes_api(bp):
    """注册已废弃端点（返回 410 Gone）"""

    @bp.route('/inventory/api/<entity>/<int:eid>', methods=['PATCH', 'DELETE'])
    def api_crud_deprecated(entity, eid):
        return jsonify({
            'ok': False,
            'msg': f'端点已废弃，请使用 routes_data.py 中的专用接口（entity={entity}, id={eid}）',
            'migrated_to': 'routes_data.py'
        }), 410

    @bp.route('/inventory/api/alert/<int:aid>/resolve', methods=['POST'])
    def alert_resolve_deprecated(aid):
        return jsonify({
            'ok': False,
            'msg': '端点已废弃',
            'migrated_to': 'routes_core.py#inventory_alert'
        }), 410

    @bp.route('/inventory/api/stock/adjust', methods=['POST'])
    def stock_adjust_deprecated():
        return jsonify({
            'ok': False,
            'msg': '端点已废弃',
            'migrated_to': 'routes_core.py#inbound_do / outbound_do'
        }), 410

    @bp.route('/inventory/api/settings', methods=['POST'])
    def save_settings_deprecated():
        """CRITICAL: 此端点曾违规写入 password 字段，已废弃
        正确实现：routes_system.py#save_settings (admin_required + 拒绝 password)"""
        return jsonify({
            'ok': False,
            'msg': '端点已废弃（曾因写 password 字段被废弃），请使用 routes_system.py',
            'migrated_to': 'routes_system.py#save_settings',
            'security_note': 'TASK-018: 拒绝通过 API 写入 password 字段'
        }), 410

    @bp.route('/inventory/api/cleanup', methods=['POST'])
    def cleanup_logs_deprecated():
        return jsonify({
            'ok': False,
            'msg': '端点已废弃（缺 admin_required 校验）',
            'migrated_to': 'routes_system.py'
        }), 410

    # ============================================================
    # TASK-T7: 报表
    # ============================================================
    @bp.route('/inventory/reports', methods=['GET'])
    @admin_required
    @require_feature('t7_reports')
    def reports_page():
        return render_template('inventory/reports.html')

    @bp.route('/inventory/api/report/stock-trend', methods=['GET'])
    @admin_required
    @require_feature('t7_reports')
    def report_stock_trend():
        from .services import ReportService
        months = int(request.args.get('months', 6))
        code, data = ReportService.stock_trend(months)
        return jsonify({'ok': code == 200, 'data': data}), code

    @bp.route('/inventory/api/report/io-flow', methods=['GET'])
    @admin_required
    @require_feature('t7_reports')
    def report_io_flow():
        from .services import ReportService
        weeks = int(request.args.get('weeks', 12))
        code, data = ReportService.inbound_outbound_flow(weeks)
        return jsonify({'ok': code == 200, 'data': data}), code

    @bp.route('/inventory/api/report/top-low-stock', methods=['GET'])
    @admin_required
    @require_feature('t7_reports')
    def report_top_low_stock():
        from .services import ReportService
        limit = int(request.args.get('limit', 10))
        code, data = ReportService.top_low_stock(limit)
        return jsonify({'ok': code == 200, 'data': data}), code

    @bp.route('/inventory/api/report/category-distribution', methods=['GET'])
    @admin_required
    @require_feature('t7_reports')
    def report_category_dist():
        from .services import ReportService
        code, data = ReportService.category_distribution()
        return jsonify({'ok': code == 200, 'data': data}), code

    # ============================================================
    # TASK-T8: 通知
    # ============================================================
    @bp.route('/inventory/notifications', methods=['GET'])
    @admin_required
    @require_feature('t8_notifications')
    def notifications_page():
        return render_template('inventory/notifications.html')

    @bp.route('/inventory/api/notification/list', methods=['GET'])
    @admin_required
    @require_feature('t8_notifications')
    def notification_list():
        from .db_utils import parse_pagination
        from .services import NotificationService
        page, page_size = parse_pagination(request.args)
        is_read = request.args.get('is_read')
        if is_read is not None:
            is_read = int(is_read)
        code, data = NotificationService.list_all(is_read, page, page_size)
        return jsonify({'ok': code == 200, **data}), code

    @bp.route('/inventory/api/notification/unread-count', methods=['GET'])
    @admin_required
    @require_feature('t8_notifications')
    def notification_unread_count():
        from .services import NotificationService
        code, items = NotificationService.list_unread(1000)
        if code == 200:
            return jsonify({'ok': True, 'count': len(items)})
        return jsonify({'ok': False, 'count': 0}), code

    @bp.route('/inventory/api/notification/<int:nid>/read', methods=['POST'])
    @admin_required
    @require_csrf
    def notification_mark_read(nid):
        from .services import NotificationService
        code, payload = NotificationService.mark_read(nid)
        return jsonify(payload), code

    @bp.route('/inventory/api/notification/read-all', methods=['POST'])
    @admin_required
    @require_csrf
    def notification_read_all():
        from .services import NotificationService
        code, payload = NotificationService.mark_all_read()
        return jsonify(payload), code

    @bp.route('/inventory/api/notification/check-low-stock', methods=['POST'])
    @admin_required
    @require_csrf
    def notification_check_low_stock():
        """手动触发低库存检查"""
        from .services import NotificationService
        count = NotificationService.auto_check_low_stock()
        return jsonify({'ok': True, 'created': count})

    # ============================================================
    # TASK-T8: 扫码录入页面
    # ============================================================
    @bp.route('/inventory/scanner', methods=['GET'])
    @admin_required
    @require_feature('t8_scanner')
    def scanner_page():
        return render_template('inventory/scanner.html')

    # ============================================================
    # TASK-T8: 导入导出（xlsx + dry-run + commit 三件套）
    # ============================================================
    @bp.route('/inventory/api/import/template', methods=['GET'])
    @admin_required
    @require_feature('t8_import_export')
    def import_template():
        """TASK-T8: 下载 xlsx 模板"""
        try:
            from openpyxl import Workbook
            from flask import send_file
            import io
        except ImportError:
            return jsonify({'ok': False, 'msg': '请安装 openpyxl: pip install openpyxl'}), 500

        entity = request.args.get('entity', 'product')
        headers = {
            'product': ['code', 'name', 'spec', 'unit', 'category_id', 'safety_stock', 'max_stock'],
            'supplier': ['name', 'contact', 'phone', 'address'],
            'category': ['name'],
            'warehouse': ['code', 'name', 'address', 'manager'],
            'base': ['code', 'name', 'address'],
        }
        if entity not in headers:
            return jsonify({'ok': False, 'msg': f'未知实体: {entity}'}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = entity
        ws.append(headers[entity])
        # 示例行
        ws.append(['示例1', '示例名称', '规格', '个', 1, 10, 100] if entity == 'product' else
                  ['示例', '', '', ''])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'{entity}_template.xlsx'
        )

    @bp.route('/inventory/api/import/dry-run', methods=['POST'])
    @admin_required
    @require_csrf
    @require_feature('t8_import_export')
    def import_dry_run():
        """TASK-T8: dry-run 不入库，仅返回错误列表"""
        from werkzeug.utils import secure_filename
        import secrets
        from datetime import datetime, timedelta
        try:
            from openpyxl import load_workbook
        except ImportError:
            return jsonify({'ok': False, 'msg': '请安装 openpyxl'}), 500

        if 'file' not in request.files:
            return jsonify({'ok': False, 'msg': '未上传文件'}), 400
        f = request.files['file']
        if not f.filename.endswith('.xlsx'):
            return jsonify({'ok': False, 'msg': '仅支持 xlsx'}), 400

        # 文件大小限制 5MB
        f.seek(0, 2)
        size = f.tell()
        f.seek(0)
        if size > 5 * 1024 * 1024:
            return jsonify({'ok': False, 'msg': '文件超过 5MB'}), 400

        entity = request.form.get('entity', 'product')
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return jsonify({'ok': False, 'msg': f'解析 xlsx 失败: {e}'}), 400

        if not rows or len(rows) < 2:
            return jsonify({'ok': False, 'msg': '文件为空'}), 400

        # 行数限制
        if len(rows) > 10000:
            return jsonify({'ok': False, 'msg': '行数超过 10000 限制'}), 400

        headers = rows[0]
        errors = []
        valid = 0
        for i, row in enumerate(rows[1:], 1):
            if not row or not any(row):
                continue
            # 必填校验
            if entity == 'product' and (not row[0] or not row[1] or not row[3]):
                errors.append({'row': i, 'msg': 'code/name/unit 不能为空'})
                continue
            valid += 1

        # 创建会话
        token = secrets.token_urlsafe(32)
        expires = (datetime.now() + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')

        execute(
            'INSERT INTO import_sessions (token, entity, file_name, file_size, total_rows, valid_rows, invalid_rows, status, error_detail, operator, expires_at) '
            'VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
            (token, entity, secure_filename(f.filename), size, len(rows) - 1, valid, len(errors),
             'pending', json.dumps(errors) if errors else None,
             session.get('username', 'admin'), expires),
            commit=True
        )

        return jsonify({
            'ok': True,
            'token': token,
            'total': len(rows) - 1,
            'valid': valid,
            'invalid': len(errors),
            'errors': errors[:50],  # 最多返回 50 条
            'expires_at': expires
        })

    @bp.route('/inventory/api/import/commit', methods=['POST'])
    @admin_required
    @require_csrf
    @require_feature('t8_import_export')
    def import_commit():
        """TASK-T8: commit - 仅 dry-run 通过才执行"""
        token = request.json.get('token') if request.is_json else None
        if not token:
            return jsonify({'ok': False, 'msg': 'token 不能为空'}), 400

        sess = execute(
            "SELECT * FROM import_sessions WHERE token=%s AND status='pending'",
            (token,), fetch='one'
        )
        if not sess:
            return jsonify({'ok': False, 'msg': '会话不存在或已使用'}), 404
        if sess['invalid_rows'] > 0:
            return jsonify({'ok': False, 'msg': f'有 {sess["invalid_rows"]} 条错误，请先修正'}), 422

        # 标记已提交
        execute(
            "UPDATE import_sessions SET status='committed', committed_at=NOW() WHERE token=%s",
            (token,), commit=True
        )

        try:
            log_operation('import_commit', sess['entity'], 0,
                          operator=session.get('username', 'admin'),
                          detail={'file': sess['file_name'], 'rows': sess['valid_rows']})
        except Exception:
            pass

        return jsonify({'ok': True, 'message': f'已标记 {sess["valid_rows"]} 条记录为待导入（实际入库请在产品/分类/供应商管理页执行）'})
