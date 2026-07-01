# -*- coding: utf-8 -*-
"""
数据导出模块 - Excel/CSV导出功能

功能说明：
- 支持导出订单、工序、库存等数据
- 支持Excel和CSV格式
- 自动处理日期、金额等字段格式化
- 支持自定义列映射和转换

使用方式：
    from data_export import export_to_excel, export_to_csv, DataExporter

    # 导出为Excel
    export_to_excel(data, filename="orders.xlsx", sheet_name="订单")

    # 导出为CSV
    export_to_csv(data, filename="orders.csv")

    # 使用Exporter类
    exporter = DataExporter()
    exporter.add_column("订单号", "order_no")
    exporter.add_column("客户", "customer_name")
    exporter.export(data, format="xlsx")
"""
import os
import io
import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Callable, Tuple
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


@dataclass
class ColumnDefinition:
    """列定义"""
    header: str
    field: str
    formatter: Optional[Callable[[Any], Any]] = None
    width: Optional[int] = None


class DataExporter:
    """
    数据导出器

    支持功能：
    - 自定义列映射
    - 数据格式化
    - Excel/CSV双格式支持
    - 大数据量分批处理

    使用示例：
        exporter = DataExporter()
        exporter.add_column("订单号", "order_no")
        exporter.add_column("金额", "total_amount", formatter=lambda x: f"¥{x:.2f}")
        exporter.add_column("创建时间", "created_at", formatter=lambda x: x.strftime("%Y-%m-%d"))

        result = exporter.export(data, format="xlsx")
    """

    def __init__(self):
        self.columns: List[ColumnDefinition] = []
        self.title: Optional[str] = None

    def set_title(self, title: str):
        """设置导出标题"""
        self.title = title
        return self

    def add_column(
        self,
        header: str,
        field: str,
        formatter: Optional[Callable[[Any], Any]] = None,
        width: Optional[int] = None
    ):
        """
        添加导出列

        参数说明：
            header (str): 列标题
            field (str): 数据字段名
            formatter (Callable): 值格式化函数
            width (int): Excel列宽
        """
        self.columns.append(ColumnDefinition(
            header=header,
            field=field,
            formatter=formatter,
            width=width
        ))
        return self

    def _get_nested_value(self, row: Dict[str, Any], field: str) -> Any:
        """获取嵌套字段值"""
        parts = field.split('.')
        value = row
        for part in parts:
            if isinstance(value, dict):
                value = value.get(part)
            else:
                return None
        return value

    def _format_value(self, value: Any, formatter: Optional[Callable]) -> Any:
        """格式化值"""
        if value is None:
            return ""

        if formatter:
            try:
                return formatter(value)
            except Exception as e:
                logger.warning(f"[DataExport] 格式化失败: {e}")
                return str(value)

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")

        if isinstance(value, (int, float)):
            return value

        return str(value)

    def _prepare_rows(self, data: List[Dict[str, Any]]) -> Tuple[List[List[Any]], List[int]]:
        """准备导出行数据"""
        headers = [col.header for col in self.columns]
        widths = [col.width or 15 for col in self.columns]

        rows = []
        for row in data:
            row_data = []
            for col in self.columns:
                value = self._get_nested_value(row, col.field)
                formatted = self._format_value(value, col.formatter)
                row_data.append(formatted)
            rows.append(row_data)

        return headers, rows, widths

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        output_path: Optional[str] = None,
        encoding: str = "utf-8-sig"
    ) -> str:
        """
        导出为CSV文件

        参数说明：
            data (List[Dict]): 要导出的数据
            output_path (str): 输出文件路径（可选）
            encoding (str): 文件编码

        返回值说明：
            str: CSV内容或文件路径
        """
        if not data:
            raise ValueError("没有数据可导出")

        headers, rows, _ = self._prepare_rows(data)

        output = io.StringIO()

        if self.title:
            output.write(f"{self.title}\n")

        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)

        content = output.getvalue()

        if output_path:
            with open(output_path, 'w', encoding=encoding, newline='') as f:
                f.write(content)
            logger.info(f"[DataExport] CSV已导出: {output_path}")
            return output_path

        return content

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        output_path: Optional[str] = None,
        sheet_name: str = "Sheet1"
    ) -> Optional[bytes]:
        """
        导出为Excel文件

        参数说明：
            data (List[Dict]): 要导出的数据
            output_path (str): 输出文件路径（可选）
            sheet_name (str): 工作表名称

        返回值说明：
            Optional[bytes]: Excel文件字节数据（如果未指定输出路径）
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.error("[DataExport] openpyxl未安装，请运行: pip install openpyxl")
            raise ImportError("需要安装openpyxl: pip install openpyxl")

        if not data:
            raise ValueError("没有数据可导出")

        headers, rows, widths = self._prepare_rows(data)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        if self.title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_col=len(headers))
            title_cell = ws.cell(row=1, column=1, value=self.title)
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            start_row = 2
        else:
            start_row = 1

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, row_data in enumerate(rows, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center")

        for col_idx, width in enumerate(widths, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        if output_path:
            wb.save(output_path)
            logger.info(f"[DataExport] Excel已导出: {output_path}")
            return None

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export(
        self,
        data: List[Dict[str, Any]],
        format: str = "xlsx",
        output_path: Optional[str] = None,
        **kwargs
    ) -> Any:
        """
        通用导出方法

        参数说明：
            data (List[Dict]): 要导出的数据
            format (str): 导出格式（xlsx/csv）
            output_path (str): 输出文件路径

        返回值说明：
            Any: 导出结果（文件路径或内容）
        """
        format = format.lower()

        if format == "xlsx":
            return self.export_to_excel(data, output_path, **kwargs)
        elif format == "csv":
            return self.export_to_csv(data, output_path, **kwargs)
        else:
            raise ValueError(f"不支持的格式: {format}")

    @classmethod
    def from_profile(cls, profile: Dict) -> "DataExporter":
        """
        从导出配置构建 DataExporter

        参数说明：
            profile (Dict): 导出配置，格式：
                {
                    'name': '报表名称',
                    'title': '表格标题',
                    'columns_config': '[{"header":"列名","field":"字段","formatter":"amount/percent"}]',
                    'sheet_name': 'Sheet1'
                }

        返回值说明：
            DataExporter: 配置好的导出器
        """
        exporter = cls()
        title = profile.get('title', '') or profile.get('name', '')
        if title:
            exporter.set_title(title)

        columns_config = profile.get('columns_config', '[]')
        if isinstance(columns_config, str):
            columns_config = json.loads(columns_config)

        for col in columns_config:
            header = col.get('header', col.get('field', ''))
            field = col.get('field', header)
            formatter_type = col.get('formatter')
            if formatter_type == 'amount':
                exporter.add_column(header, field, formatter=lambda x: f"¥{float(x):.2f}" if x else "¥0.00")
            elif formatter_type == 'percent':
                exporter.add_column(header, field, formatter=lambda x: f"{float(x):.1f}%" if isinstance(x, (int, float)) else str(x))
            else:
                exporter.add_column(header, field)

        return exporter


def export_to_csv(
    data: List[Dict[str, Any]],
    filename: str,
    columns: Optional[List[Tuple[str, str]]] = None,
    output_dir: Optional[str] = None
) -> str:
    """
    快速导出CSV

    参数说明：
        data (List[Dict]): 要导出的数据
        filename (str): 文件名
        columns (List[Tuple]): 列定义 [(标题, 字段名), ...]
        output_dir (str): 输出目录

    返回值说明：
        str: 导出文件路径
    """
    exporter = DataExporter()

    if columns:
        for header, field in columns:
            exporter.add_column(header, field)
    elif data:
        first_row = data[0]
        for field in first_row.keys():
            exporter.add_column(field, field)

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
    else:
        output_path = filename

    return exporter.export_to_csv(data, output_path)


def export_to_excel(
    data: List[Dict[str, Any]],
    filename: str,
    sheet_name: str = "数据",
    columns: Optional[List[Tuple[str, str]]] = None,
    output_dir: Optional[str] = None,
    title: Optional[str] = None
) -> str:
    """
    快速导出Excel

    参数说明：
        data (List[Dict]): 要导出的数据
        filename (str): 文件名
        sheet_name (str): 工作表名称
        columns (List[Tuple]): 列定义 [(标题, 字段名), ...]
        output_dir (str): 输出目录
        title (str): 表格标题

    返回值说明：
        str: 导出文件路径
    """
    exporter = DataExporter()

    if title:
        exporter.set_title(title)

    if columns:
        for header, field in columns:
            exporter.add_column(header, field)
    elif data:
        first_row = data[0]
        for field in first_row.keys():
            exporter.add_column(field, field)

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
    else:
        output_path = filename

    exporter.export_to_excel(data, output_path, sheet_name)
    return output_path


def export_orders(orders: List[Dict[str, Any]], format: str = "xlsx") -> Any:
    """
    导出订单数据

    参数说明：
        orders (List[Dict]): 订单数据列表
        format (str): 导出格式

    返回值说明：
        Any: 导出结果
    """
    exporter = DataExporter()
    exporter.set_title("订单导出")

    exporter.add_column("订单号", "order_no")
    exporter.add_column("客户", "customer_name")
    exporter.add_column("产品", "product_name")
    exporter.add_column("规格", "spec")
    exporter.add_column("数量", "quantity")
    exporter.add_column("单位", "unit")
    exporter.add_column("单价", "unit_price", formatter=lambda x: f"¥{x:.2f}" if x else "¥0.00")
    exporter.add_column("金额", "total_amount", formatter=lambda x: f"¥{x:.2f}" if x else "¥0.00")
    exporter.add_column("状态", "status")
    exporter.add_column("创建时间", "created_at",
                        formatter=lambda x: x.strftime("%Y-%m-%d %H:%M") if x else "")

    return exporter.export(orders, format=format)


def export_process_records(records: List[Dict[str, Any]], format: str = "xlsx") -> Any:
    """
    导出工序记录

    参数说明：
        records (List[Dict]): 工序记录数据列表
        format (str): 导出格式

    返回值说明：
        Any: 导出结果
    """
    exporter = DataExporter()
    exporter.set_title("工序记录导出")

    exporter.add_column("订单号", "order_no")
    exporter.add_column("工序", "process_name")
    exporter.add_column("操作员", "operator_name")
    exporter.add_column("数量", "quantity")
    exporter.add_column("合格数", "qualified_count")
    exporter.add_column("不合格数", "rejected_count")
    exporter.add_column("开始时间", "start_time",
                        formatter=lambda x: x.strftime("%Y-%m-%d %H:%M") if x else "")
    exporter.add_column("结束时间", "end_time",
                        formatter=lambda x: x.strftime("%Y-%m-%d %H:%M") if x else "")
    exporter.add_column("耗时(分钟)", "duration_minutes")

    return exporter.export(records, format=format)


def export_inventory(inventory: List[Dict[str, Any]], format: str = "xlsx") -> Any:
    """
    导出库存数据

    参数说明：
        inventory (List[Dict]): 库存数据列表
        format (str): 导出格式

    返回值说明：
        Any: 导出结果
    """
    exporter = DataExporter()
    exporter.set_title("库存导出")

    exporter.add_column("物料编码", "material_code")
    exporter.add_column("物料名称", "material_name")
    exporter.add_column("规格", "spec")
    exporter.add_column("数量", "quantity")
    exporter.add_column("单位", "unit")
    exporter.add_column("仓库", "warehouse")
    exporter.add_column("库位", "location")
    exporter.add_column("更新时间", "updated_at",
                        formatter=lambda x: x.strftime("%Y-%m-%d %H:%M") if x else "")

    return exporter.export(inventory, format=format)


def export_order_costs(costs: List[Dict[str, Any]], format: str = "xlsx") -> Any:
    """
    导出订单成本核算数据

    参数说明：
        costs (List[Dict]): 成本数据列表
        format (str): 导出格式

    返回值说明：
        Any: 导出结果
    """
    exporter = DataExporter()
    exporter.set_title("订单成本核算")

    exporter.add_column("订单号", "order_no")
    exporter.add_column("客户", "customer_name")
    exporter.add_column("产品", "product_name")
    exporter.add_column("数量", "quantity")
    exporter.add_column("单位", "unit")
    exporter.add_column("收入", "revenue", formatter=lambda x: f"{x:.2f}" if x else "0.00")
    exporter.add_column("材料成本", "material_cost", formatter=lambda x: f"{x:.2f}" if x else "0.00")
    exporter.add_column("人工成本", "labor_cost", formatter=lambda x: f"{x:.2f}" if x else "0.00")
    exporter.add_column("制造费用", "overhead_cost", formatter=lambda x: f"{x:.2f}" if x else "0.00")
    exporter.add_column("外协成本", "outsourcing_cost", formatter=lambda x: f"{x:.2f}" if x else "0.00")
    exporter.add_column("其他成本", "other_cost", formatter=lambda x: f"{x:.2f}" if x else "0.00")
    exporter.add_column("总成本", "total_cost", formatter=lambda x: f"{x:.2f}" if x else "0.00")
    exporter.add_column("利润", "profit", formatter=lambda x: f"{x:.2f}" if x else "0.00")
    exporter.add_column("利润率", "margin_rate", formatter=lambda x: f"{x:.1f}%" if x else "0.0%")
    exporter.add_column("核算时间", "calculated_at",
                        formatter=lambda x: x if x else "")

    return exporter.export(costs, format=format)
