# -*- coding: utf-8 -*-
"""
库存管理系统备份模块
支持数据库备份、恢复、定时任务
"""
import os
import sys
import shutil
import time
import subprocess
import threading
from datetime import datetime, date

try:
    from inventory_db_complete import INVENTORY_DB_CONFIG
except ImportError:
    from inventory_mysql_db import INVENTORY_DB_CONFIG


BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "inventory_backups")
BACKUP_KEEP_DAYS = 30


def get_backup_dir():
    """获取备份目录"""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    return BACKUP_DIR


def get_backup_files():
    """获取所有备份文件列表"""
    backup_files = []
    for filename in os.listdir(get_backup_dir()):
        if filename.endswith('.sql'):
            filepath = os.path.join(get_backup_dir(), filename)
            stat = os.stat(filepath)
            backup_files.append({
                'filename': filename,
                'filepath': filepath,
                'size': stat.st_size,
                'created': datetime.fromtimestamp(stat.st_ctime),
                'modified': datetime.fromtimestamp(stat.st_mtime)
            })
    return sorted(backup_files, key=lambda x: x['modified'], reverse=True)


def backup_database():
    """备份库存系统数据库"""
    return backup_database_fallback()


def backup_database_fallback():
    """使用Python方式备份数据库（当mysqldump不可用时）"""
    import pymysql
    from pymysql.cursors import DictCursor

    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(get_backup_dir(), f"inventory_backup_{timestamp}.sql")

        config = INVENTORY_DB_CONFIG.copy()
        db_name = config['database']

        conn = pymysql.connect(
            host=config['host'],
            port=config['port'],
            user=config['user'],
            password=config['password'],
            database=db_name,
            charset='utf8mb4',
            cursorclass=DictCursor
        )
        cursor = conn.cursor()

        cursor.execute("SHOW TABLES")
        tables = [list(row.values())[0] for row in cursor.fetchall()]

        with open(backup_file, 'w', encoding='utf-8') as f:
            f.write(f"-- 库存管理系统数据库备份\n")
            f.write(f"-- 数据库: {db_name}\n")
            f.write(f"-- 备份时间: {timestamp}\n")
            f.write(f"-- 版本: V3.0 MySQL版\n\n")

            for table in tables:
                cursor.execute(f"SHOW CREATE TABLE `{table}`")
                create_result = cursor.fetchone()
                create_stmt = list(create_result.values())[0]
                f.write(f"\n-- 表结构: {table}\n")
                f.write(f"DROP TABLE IF EXISTS `{table}`;\n")
                f.write(f"{create_stmt};\n\n")

                cursor.execute(f"SELECT * FROM `{table}`")
                rows = cursor.fetchall()

                if rows:
                    cursor.execute(f"SHOW COLUMNS FROM `{table}`")
                    columns = [col['Field'] for col in cursor.fetchall()]

                    f.write(f"-- 数据: {table} ({len(rows)} rows)\n")
                    for row in rows:
                        values = []
                        for col in columns:
                            val = row[col]
                            if val is None:
                                values.append('NULL')
                            elif isinstance(val, (int, float)):
                                values.append(str(val))
                            elif isinstance(val, datetime):
                                values.append(f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'")
                            elif isinstance(val, date):
                                values.append(f"'{val.strftime('%Y-%m-%d')}'")
                            else:
                                val_str = str(val)
                                val_str = val_str.replace('\\', '\\\\').replace("'", "\\'").replace('"', '\\"')
                                values.append(f"'{val_str}'")
                        f.write(f"INSERT INTO `{table}` ({', '.join([f'`{c}`' for c in columns])}) VALUES ({', '.join(values)});\n")
                f.write("\n")

        cursor.close()
        conn.close()

        file_size = os.path.getsize(backup_file)
        print(f"[备份] 数据库备份成功(Python方式): {backup_file} ({file_size} bytes)")

        cleanup_old_backups(BACKUP_KEEP_DAYS)

        return {
            'success': True,
            'filename': os.path.basename(backup_file),
            'filepath': backup_file,
            'size': file_size,
            'message': f'备份成功: {os.path.basename(backup_file)}'
        }

    except Exception as e:
        print(f"[备份] Python方式备份失败: {e}")
        return {'success': False, 'message': f'备份失败: {str(e)}'}


def restore_database(backup_filepath):
    """恢复数据库"""
    try:
        if not os.path.exists(backup_filepath):
            return {'success': False, 'message': '备份文件不存在'}

        config = INVENTORY_DB_CONFIG.copy()
        db_name = config.pop('database')

        try:
            cmd = [
                'mysql',
                f"--host={config['host']}",
                f"--port={config['port']}",
                f"--user={config['user']}",
                f"--password={config['password']}",
                db_name
            ]

            with open(backup_filepath, 'r', encoding='utf-8') as f:
                result = subprocess.run(cmd, stdin=f, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

            if result.returncode != 0:
                raise Exception(f"mysql restore failed: {result.stderr}")

        except FileNotFoundError:
            restore_database_fallback(backup_filepath)

        print(f"[备份] 数据库恢复成功: {backup_filepath}")
        return {'success': True, 'message': '数据库恢复成功'}

    except Exception as e:
        print(f"[备份] 数据库恢复失败: {e}")
        return {'success': False, 'message': f'恢复失败: {str(e)}'}


def restore_database_fallback(backup_filepath):
    """使用Python方式恢复数据库"""
    import pymysql

    config = INVENTORY_DB_CONFIG.copy()
    db_name = config['database']

    conn = pymysql.connect(
        host=config['host'],
        port=config['port'],
        user=config['user'],
        password=config['password'],
        charset='utf8mb4'
    )
    cursor = conn.cursor()

    with open(backup_filepath, 'r', encoding='utf-8') as f:
        sql_content = f.read()

    statements = []
    current_stmt = []
    in_string = False
    escape_next = False

    for char in sql_content:
        if escape_next:
            current_stmt.append(char)
            escape_next = False
            continue
        if char == '\\':
            escape_next = True
            current_stmt.append(char)
            continue
        if char == "'" and not escape_next:
            in_string = not in_string
        if char == ';' and not in_string:
            stmt = ''.join(current_stmt).strip()
            if stmt and not stmt.startswith('--'):
                statements.append(stmt)
            current_stmt = []
        else:
            current_stmt.append(char)

    cursor.execute(f"DROP DATABASE IF EXISTS {db_name}")
    cursor.execute(f"CREATE DATABASE {db_name} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
    cursor.execute(f"USE {db_name}")

    for stmt in statements:
        if stmt.strip():
            try:
                cursor.execute(stmt)
            except Exception as e:
                print(f"[警告] 执行SQL失败: {stmt[:100]}... 错误: {e}")

    conn.commit()
    cursor.close()
    conn.close()


def cleanup_old_backups(days=30):
    """清理旧备份文件"""
    cutoff_time = time.time() - (days * 24 * 60 * 60)
    backup_dir = get_backup_dir()

    for filename in os.listdir(backup_dir):
        if filename.endswith('.sql'):
            filepath = os.path.join(backup_dir, filename)
            if os.path.isfile(filepath) and os.path.getmtime(filepath) < cutoff_time:
                try:
                    os.remove(filepath)
                    print(f"[备份] 已删除过期备份: {filename}")
                except Exception as e:
                    print(f"[备份] 删除旧备份失败 {filename}: {e}")


def export_to_excel():
    """导出数据到Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return {'success': False, 'message': 'openpyxl模块未安装'}

    try:
        import pymysql
        from pymysql.cursors import DictCursor

        config = INVENTORY_DB_CONFIG.copy()
        db_name = config['database']

        conn = pymysql.connect(
            host=config['host'],
            port=config['port'],
            user=config['user'],
            password=config['password'],
            database=db_name,
            charset='utf8mb4',
            cursorclass=DictCursor
        )
        cursor = conn.cursor()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(get_backup_dir(), f"inventory_export_{timestamp}.xlsx")

        wb = Workbook()
        wb.remove(wb.active)

        style_font = Font(name='微软雅黑', size=10)
        style_header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        style_header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        style_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        tables = ['products', 'inventory', 'warehouses', 'categories', 'suppliers', 'inventory_transactions']
        table_names_cn = {
            'products': '商品信息',
            'inventory': '库存台账',
            'warehouses': '仓库信息',
            'categories': '商品分类',
            'suppliers': '供应商信息',
            'inventory_transactions': '库存流水'
        }

        for table in tables:
            try:
                cursor.execute(f"SELECT * FROM {table}")
                rows = cursor.fetchall()

                if not rows:
                    continue

                ws = wb.create_sheet(title=table_names_cn.get(table, table))

                headers = list(rows[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = style_header_font
                    cell.fill = style_header_fill
                    cell.alignment = style_alignment
                    cell.border = thin_border

                for row_idx, row in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        val = row[header]
                        if isinstance(val, datetime):
                            val = val.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(val, date):
                            val = val.strftime('%Y-%m-%d')
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.font = style_font
                        cell.border = thin_border
                        cell.alignment = style_alignment

                for col_idx in range(1, len(headers) + 1):
                    ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 15

            except Exception as e:
                print(f"[导出] 导出表 {table} 失败: {e}")
                continue

        wb.save(excel_file)
        cursor.close()
        conn.close()

        file_size = os.path.getsize(excel_file)
        print(f"[导出] 数据导出成功: {excel_file} ({file_size} bytes)")

        return {
            'success': True,
            'filename': os.path.basename(excel_file),
            'filepath': excel_file,
            'size': file_size,
            'message': f'导出成功: {os.path.basename(excel_file)}'
        }

    except Exception as e:
        print(f"[导出] 数据导出失败: {e}")
        return {'success': False, 'message': f'导出失败: {str(e)}'}


def auto_backup_loop():
    """自动备份循环（每天凌晨2点执行）"""
    while True:
        try:
            now = datetime.now()
            next_backup_hour = 2

            next_backup = datetime(now.year, now.month, now.day, next_backup_hour, 0, 0)
            if now.hour >= next_backup_hour:
                next_backup = next_backup.replace(day=next_backup.day + 1)

            sleep_seconds = (next_backup - now).total_seconds()
            sleep_seconds = min(sleep_seconds, 86400)

            time.sleep(sleep_seconds)

            if datetime.now().hour == next_backup_hour:
                print("[自动备份] 开始执行每日备份...")
                backup_database()

        except Exception as e:
            print(f"[自动备份] 自动备份出错: {e}")
            time.sleep(3600)


def start_auto_backup():
    """启动自动备份线程"""
    backup_thread = threading.Thread(target=auto_backup_loop, daemon=True)
    backup_thread.start()
    print("[备份系统] 自动备份服务已启动（每天凌晨2点执行）")
    return backup_thread


if __name__ == "__main__":
    print("=" * 50)
    print("库存管理系统 备份工具 V3.0")
    print("=" * 50)
    print()

    while True:
        print("\n选项:")
        print("1. 立即备份数据库")
        print("2. 列出所有备份")
        print("3. 恢复数据库")
        print("4. 导出数据到Excel")
        print("5. 清理旧备份")
        print("0. 退出")
        print()

        choice = input("请选择操作: ").strip()

        if choice == '1':
            print("\n正在备份数据库...")
            result = backup_database()
            print(f"结果: {result['message']}")

        elif choice == '2':
            print("\n备份文件列表:")
            files = get_backup_files()
            if not files:
                print("  暂无备份文件")
            else:
                for i, f in enumerate(files, 1):
                    print(f"  {i}. {f['filename']}")
                    print(f"     大小: {f['size']:,} bytes")
                    print(f"     创建时间: {f['created'].strftime('%Y-%m-%d %H:%M:%S')}")

        elif choice == '3':
            files = get_backup_files()
            if not files:
                print("\n暂无备份文件可恢复")
                continue

            print("\n请选择要恢复的备份文件编号:")
            for i, f in enumerate(files, 1):
                print(f"  {i}. {f['filename']} ({f['created'].strftime('%Y-%m-%d %H:%M:%S')})")

            try:
                idx = int(input("\n编号: ").strip()) - 1
                if 0 <= idx < len(files):
                    confirm = input(f"确定要恢复数据库吗？这将覆盖当前数据！ (yes/no): ")
                    if confirm.lower() == 'yes':
                        result = restore_database(files[idx]['filepath'])
                        print(f"结果: {result['message']}")
            except ValueError:
                print("无效的编号")

        elif choice == '4':
            print("\n正在导出数据到Excel...")
            result = export_to_excel()
            print(f"结果: {result.get('message', '导出失败')}")

        elif choice == '5':
            print("\n正在清理30天前的备份...")
            cleanup_old_backups(30)
            print("清理完成")

        elif choice == '0':
            print("\n感谢使用！")
            break
