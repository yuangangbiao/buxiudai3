# -*- coding: utf-8 -*-
"""
导入订单预览确认弹窗（v3.3）

v3.3 修定:
- 列名匹配代替位置索引（map_headers）
- Treeview 增加"订单号"列（空显示"(自动生成)"）
"""
import os
import tkinter as tk
from tkinter import ttk, messagebox
from config import COLORS, FONTS, LAYOUT
from desktop.views.dialogs import center_window
from utils.excel_utils import map_headers, _to_int, _to_float


class ImportPreviewDialog(tk.Toplevel):
    """导入订单预览确认弹窗"""

    def __init__(self, parent, file_path, on_confirm):
        super().__init__(parent)
        self.file_path = file_path
        self.on_confirm = on_confirm
        self.preview_data = []
        self.total_rows = 0
        self.errors = []

        self.title("导入预览确认")
        self.geometry("800x500")
        center_window(self, 800, 500)
        self.resizable(True, True)

        # 不允许重复打开
        self.transient(parent)
        self.grab_set()

        self._build_ui()
        self._load_preview()

    def _build_ui(self):
        """构建界面"""
        # 标题区
        header = tk.Frame(self, bg=COLORS["primary"], height=40)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(header, text="📋 导入订单预览",
                font=FONTS["subtitle"], bg=COLORS["primary"], fg="white").pack(side=tk.LEFT, padx=15, pady=8)

        # 文件信息
        info_frame = tk.Frame(self, bg=COLORS["bg_card"], padx=15, pady=10)
        info_frame.pack(fill=tk.X)

        file_name = os.path.basename(self.file_path)
        tk.Label(info_frame, text=f"文件: {file_name}",
                font=FONTS["body"], bg=COLORS["bg_card"]).pack(anchor="w")
        self.info_label = tk.Label(info_frame, text="正在读取...",
                font=FONTS["small"], bg=COLORS["bg_card"], fg="#666666")
        self.info_label.pack(anchor="w", pady=(2, 0))

        # 预览表格区
        table_frame = tk.Frame(self, bg=COLORS["bg_card"], padx=10, pady=5)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))

        # 列定义（v3.3 增加"订单号"列）
        columns = ("order_no", "customer", "product", "material", "qty", "unit", "amount", "delivery", "status")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=12)

        self.tree.heading("order_no", text="订单号")
        self.tree.heading("customer", text="客户名称")
        self.tree.heading("product", text="产品类型")
        self.tree.heading("material", text="材质")
        self.tree.heading("qty", text="数量")
        self.tree.heading("unit", text="单位")
        self.tree.heading("amount", text="总价")
        self.tree.heading("delivery", text="交货日期")
        self.tree.heading("status", text="状态")

        self.tree.column("order_no", width=120, anchor="center")
        self.tree.column("customer", width=120, anchor="w")
        self.tree.column("product", width=100, anchor="w")
        self.tree.column("material", width=80, anchor="w")
        self.tree.column("qty", width=60, anchor="center")
        self.tree.column("unit", width=50, anchor="center")
        self.tree.column("amount", width=80, anchor="e")
        self.tree.column("delivery", width=90, anchor="center")
        self.tree.column("status", width=70, anchor="center")

        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview", font=FONTS["body"], rowheight=28)

        # 按钮区
        btn_frame = tk.Frame(self, bg=COLORS["bg_card"])
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        # 警告提示
        warn_frame = tk.Frame(btn_frame, bg="#FFF3CD", padx=10, pady=5)
        warn_frame.pack(side=tk.LEFT, padx=(0, 20))
        tk.Label(warn_frame, text="⚠️ 请确认以上订单信息无误后再导入",
                font=FONTS["small"], bg="#FFF3CD", fg="#856404").pack()

        # 操作按钮
        btn_right = tk.Frame(btn_frame, bg=COLORS["bg_card"])
        btn_right.pack(side=tk.RIGHT)

        ttk.Button(btn_right, text="取消", command=self.destroy, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_right, text="确认导入", command=self._do_confirm,
                  style="Accent.TButton", width=12).pack(side=tk.LEFT, padx=5)

    def _load_preview(self):
        """加载预览数据（v3.3: 列名匹配）"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active

            # 获取表头
            headers = [cell.value for cell in ws[1]]

            # 列名匹配（map_headers，含"订单号"→"order_no"，v3.3）
            col_map = map_headers(headers)

            # 计算总行数（排除表头）
            self.total_rows = ws.max_row - 1 if ws.max_row > 1 else 0

            # 读取前5行预览数据（列名匹配，v3.3）
            preview_count = 0
            for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
                # 跳过空行（检查所有 col_map 中的字段）
                if not any(col_idx < len(row) and row[col_idx] is not None and str(row[col_idx]).strip()
                           for col_idx in col_map.keys()):
                    continue
                preview_count += 1

                # 列名匹配提取（v3.3 D2 预览空显示"(自动生成)"）
                def get_val(field_key, default="-"):
                    for col_idx, fk in col_map.items():
                        if fk == field_key and col_idx < len(row):
                            v = row[col_idx]
                            if v is not None and str(v).strip():
                                return str(v).strip()
                            return default
                    return default

                order_no_raw = None
                for col_idx, fk in col_map.items():
                    if fk == "order_no" and col_idx < len(row):
                        v = row[col_idx]
                        if v is not None and str(v).strip():
                            order_no_raw = str(v).strip()

                customer = get_val("customer_name")
                product = get_val("product_type")
                material = get_val("material")
                qty_raw = None
                for col_idx, fk in col_map.items():
                    if fk == "quantity" and col_idx < len(row):
                        qty_raw = row[col_idx]
                qty = _to_int(qty_raw) if qty_raw else 0
                unit = get_val("unit", "米")
                amount_raw = None
                for col_idx, fk in col_map.items():
                    if fk == "total_amount" and col_idx < len(row):
                        amount_raw = row[col_idx]
                amount = _to_float(amount_raw) if amount_raw else 0
                delivery = get_val("delivery_date", "-")
                if delivery != "-" and len(delivery) > 10:
                    delivery = delivery[:10]
                status = get_val("status", "待确认")

                self.preview_data.append({
                    "order_no": order_no_raw or "(自动生成)",
                    "customer": customer,
                    "product": product,
                    "material": material,
                    "qty": qty,
                    "unit": unit,
                    "amount": amount,
                    "delivery": delivery,
                    "status": status,
                })

            wb.close()

            # 更新显示
            for data in self.preview_data:
                self.tree.insert("", tk.END, values=(
                    data["order_no"],
                    data["customer"],
                    data["product"],
                    data["material"],
                    data["qty"],
                    data["unit"],
                    f"{data['amount']:.2f}",
                    data["delivery"],
                    data["status"],
                ))

            # 更新信息标签
            if self.total_rows > 0:
                preview_msg = f"共 {self.total_rows} 条订单，将导入以下前 {len(self.preview_data)} 条作为预览"
                if self.total_rows > len(self.preview_data):
                    preview_msg += f"（共 {self.total_rows} 条，仅显示前 {len(self.preview_data)} 条）"
                self.info_label.config(text=preview_msg)
            else:
                self.info_label.config(text="⚠️ Excel文件中没有找到有效订单数据", fg="red")

        except Exception as e:
            self.errors.append(str(e))
            self.info_label.config(text=f"读取Excel文件失败: {str(e)}", fg="red")

    def _do_confirm(self):
        """确认导入"""
        if self.total_rows == 0:
            messagebox.showwarning("无可导入数据", "Excel文件中没有找到有效订单数据！", parent=self)
            return

        # 确认提示
        if messagebox.askyesno("确认导入",
                f"即将导入 {self.total_rows} 条订单。\n\n确认无误后点击「是」开始导入。",
                parent=self):
            self.destroy()
            if self.on_confirm:
                self.on_confirm()
