# -*- coding: utf-8 -*-
"""utils/excel_service.py 的单元测试

覆盖模块:
- _format_val / _format_dt: 值格式化
- _parse_date / _parse_num: 字符串解析
- _normalize_header / _normalize_operator_header: 表头归一化
- build_order_export / build_shipment_export / build_workreport_export / build_operator_export: 导出
- generate_template / generate_operator_template: 模板生成
- parse_order_import / parse_operator_import: 导入解析
"""
import os
import sys
import tempfile
from datetime import datetime, date
from io import BytesIO

import pytest
from openpyxl import load_workbook

# 把项目根目录加入 sys.path
from utils.excel_service import (
    _format_val, _format_dt, _parse_date, _parse_num,
    _normalize_header, _normalize_operator_header,
    build_order_export, build_shipment_export, build_workreport_export,
    build_operator_export, generate_template, generate_operator_template,
    parse_order_import, parse_operator_import,
    REQUIRED_FIELDS, FIELD_ALIASES, OPERATOR_REQUIRED_FIELDS, OPERATOR_FIELD_ALIASES,
)


# ── 私有函数测试 ──

class TestFormatVal:
    """_format_val 测试"""

    def test_none_returns_empty_string(self):
        assert _format_val(None) == ''

    def test_datetime_formatted(self):
        dt = datetime(2026, 6, 24, 10, 30, 0)
        assert _format_val(dt) == '2026-06-24'

    def test_date_formatted(self):
        d = date(2026, 1, 1)
        assert _format_val(d) == '2026-01-01'

    def test_string_passthrough(self):
        assert _format_val('hello') == 'hello'

    def test_number_passthrough(self):
        assert _format_val(42) == 42
        assert _format_val(3.14) == 3.14

    def test_zero_passthrough(self):
        """0 不应被格式化为空字符串"""
        assert _format_val(0) == 0


class TestFormatDt:
    """_format_dt 测试"""

    def test_none_returns_empty(self):
        assert _format_dt(None) == ''

    def test_empty_string_returns_empty(self):
        assert _format_dt('') == ''

    def test_datetime_formatted(self):
        dt = datetime(2026, 6, 24, 10, 30, 45)
        assert _format_dt(dt) == '2026-06-24 10:30:45'

    def test_string_truncated_to_19(self):
        """字符串截断到 19 字符"""
        s = '2026-06-24 10:30:45.123456 extra'
        assert _format_dt(s) == '2026-06-24 10:30:45'


class TestParseDate:
    """_parse_date 测试"""

    def test_none_returns_none(self):
        assert _parse_date(None) is None

    def test_empty_string_returns_none(self):
        assert _parse_date('') is None

    def test_datetime_obj(self):
        dt = datetime(2026, 6, 24)
        assert _parse_date(dt) == '2026-06-24'

    def test_date_obj(self):
        d = date(2026, 1, 1)
        assert _parse_date(d) == '2026-01-01'

    def test_dash_format(self):
        assert _parse_date('2026-06-24') == '2026-06-24'

    def test_slash_format(self):
        assert _parse_date('2026/06/24') == '2026-06-24'

    def test_dot_format(self):
        assert _parse_date('2026.06.24') == '2026-06-24'

    def test_compact_format(self):
        assert _parse_date('20260624') == '2026-06-24'

    def test_invalid_returns_original(self):
        """无法解析时返回原字符串"""
        assert _parse_date('not-a-date') == 'not-a-date'

    def test_whitespace_stripped(self):
        assert _parse_date('  2026-06-24  ') == '2026-06-24'


class TestParseNum:
    """_parse_num 测试"""

    def test_none_returns_none(self):
        assert _parse_num(None) is None

    def test_empty_string_returns_none(self):
        assert _parse_num('') is None

    def test_int_passthrough(self):
        assert _parse_num(42) == 42

    def test_float_passthrough(self):
        assert _parse_num(3.14) == 3.14

    def test_string_int_to_float(self):
        assert _parse_num('100') == 100.0

    def test_string_float(self):
        assert _parse_num('3.14') == 3.14

    def test_invalid_string_returns_original(self):
        """无法解析时返回原字符串"""
        assert _parse_num('abc') == 'abc'

    def test_whitespace_stripped(self):
        assert _parse_num('  42  ') == 42.0


class TestNormalizeHeader:
    """_normalize_header 测试"""

    def test_exact_match(self):
        assert _normalize_header('客户名称') == 'customer_name'
        assert _normalize_header('订单号') == 'order_no'

    def test_alias_match(self):
        assert _normalize_header('客户') == 'customer_name'
        assert _normalize_header('产品') == 'product_type'
        assert _normalize_header('qty') == 'quantity'

    def test_contains_match(self):
        """包含匹配：'网孔尺寸' 包含 '网孔尺寸'，匹配"""
        assert _normalize_header('网孔尺寸') == 'mesh_size'

    def test_unknown_returns_none(self):
        assert _normalize_header('未知字段') is None

    def test_strip_whitespace(self):
        assert _normalize_header('  客户名称  ') == 'customer_name'

    def test_required_fields(self):
        """REQUIRED_FIELDS 完整性"""
        assert 'customer_name' in REQUIRED_FIELDS
        assert 'product_type' in REQUIRED_FIELDS
        assert 'quantity' in REQUIRED_FIELDS


class TestNormalizeOperatorHeader:
    """_normalize_operator_header 测试"""

    def test_exact_match(self):
        assert _normalize_operator_header('姓名') == 'name'
        assert _normalize_operator_header('操作员ID') == 'operator_id'

    def test_alias_match(self):
        assert _normalize_operator_header('工号') == 'operator_id'
        assert _normalize_operator_header('微信号') == 'wechat_userid'

    def test_unknown_returns_none(self):
        assert _normalize_operator_header('未知') is None

    def test_required_fields(self):
        assert OPERATOR_REQUIRED_FIELDS == ['name']


# ── 导出函数测试 ──

class TestBuildOrderExport:
    """build_order_export 测试"""

    def test_empty_orders(self, tmp_path, monkeypatch):
        """空订单列表也能生成（只有表头）"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        out = build_order_export([], 'empty_orders.xlsx')
        assert os.path.exists(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == '订单列表'
        # 表头 21 列
        assert ws.cell(1, 1).value == '订单号'
        assert ws.cell(1, 21).value == '创建时间'

    def test_with_full_order(self, tmp_path, monkeypatch):
        """完整订单数据"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        orders = [{
            'order_no': 'ORD-001',
            'product_type': '普通网带',
            'customer_name': '示例客户',
            'customer_group': '华东',
            'material': '304',
            'mesh_size': '10mm×10mm',
            'wire_diameter': 2.0,
            'width': 1000,
            'length': 50,
            'quantity': 100,
            'unit': '米',
            'unit_price': 50.0,
            'total_amount': 5000.0,
            'surface_treatment': '抛光',
            'special_requirements': '无',
            'delivery_date': '2026-12-31',
            'order_status': '待确认',
            'remark': '加急',
            'customer_phone': '13800138000',
            'customer_address': '示例地址',
            'created_at': '2026-06-24 10:00:00',
        }]
        out = build_order_export(orders, 'full_order.xlsx')
        assert os.path.exists(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 1).value == 'ORD-001'
        assert ws.cell(2, 3).value == '示例客户'
        assert ws.cell(2, 10).value == 100  # quantity

    def test_extra_params_dict(self, tmp_path, monkeypatch):
        """extra_params 为 dict 时不抛错"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        orders = [{
            'order_no': 'ORD-002',
            'extra_params': {'mesh_size': '5mm'},
        }]
        out = build_order_export(orders, 'extra_dict.xlsx')
        assert os.path.exists(out)

    def test_extra_params_json_string(self, tmp_path, monkeypatch):
        """extra_params 为 JSON 字符串时正常解析"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        orders = [{
            'order_no': 'ORD-003',
            'extra_params': '{"mesh_size": "5mm"}',
        }]
        out = build_order_export(orders, 'extra_json.xlsx')
        assert os.path.exists(out)

    def test_extra_params_invalid_json(self, tmp_path, monkeypatch):
        """extra_params 为非法 JSON 时降级为空 dict"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        orders = [{
            'order_no': 'ORD-004',
            'extra_params': 'not-valid-json{',
        }]
        out = build_order_export(orders, 'extra_bad.xlsx')
        assert os.path.exists(out)

    def test_filename_sanitized(self, tmp_path, monkeypatch):
        """文件名非法字符被剥离"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        orders = [{'order_no': 'X'}]
        # 包含 / \ : * ? " < > | 字符
        out = build_order_export(orders, 'a/b\\c:d*e?f"g<h>i|j.xlsx')
        # 应安全写入（不带非法字符）
        assert os.path.exists(out)
        # 文件名应不含非法字符
        basename = os.path.basename(out)
        for ch in ('/', '\\', ':', '*', '?', '"', '<', '>', '|'):
            assert ch not in basename

    def test_product_name_fallback(self, tmp_path, monkeypatch):
        """product_type 为空时回退到 product_name"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        orders = [{
            'order_no': 'ORD-005',
            'product_name': '备用产品名',
            # product_type 缺失
        }]
        out = build_order_export(orders, 'fallback.xlsx')
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 2).value == '备用产品名'

    def test_total_amount_fallback(self, tmp_path, monkeypatch):
        """total_amount 为空时回退到 total"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        orders = [{
            'order_no': 'ORD-006',
            'total': 9999,
            # total_amount 缺失
        }]
        out = build_order_export(orders, 'total.xlsx')
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 13).value == 9999

    def test_status_fallback(self, tmp_path, monkeypatch):
        """order_status 为空时回退到 status"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        orders = [{
            'order_no': 'ORD-007',
            'status': '已完成',
        }]
        out = build_order_export(orders, 'status.xlsx')
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 17).value == '已完成'


class TestBuildShipmentExport:
    """build_shipment_export 测试"""

    def test_empty_records(self, tmp_path, monkeypatch):
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        out = build_shipment_export([], 'empty_ship.xlsx')
        assert os.path.exists(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == '发货记录'
        assert ws.cell(1, 1).value == '发货单号'

    def test_with_records(self, tmp_path, monkeypatch):
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        records = [{
            'shipment_no': 'SH-001',
            'order_no': 'ORD-001',
            'customer_name': '客户A',
            'product_type': '网带',
            'quantity': 100,
            'unit': '米',
            'ship_date': '2026-06-24',
            'receiver': '张三',
            'receiver_address': '上海',
            'logistics_company': '顺丰',
            'tracking_no': 'SF123456',
            'shipper': '李四',
            'status': '已发货',
            'created_at': '2026-06-24 09:00:00',
        }]
        out = build_shipment_export(records, 'ship.xlsx')
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 1).value == 'SH-001'
        assert ws.cell(2, 5).value == 100


class TestBuildWorkreportExport:
    """build_workreport_export 测试"""

    def test_empty(self, tmp_path, monkeypatch):
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        out = build_workreport_export([], 'empty_wr.xlsx')
        assert os.path.exists(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == '报工记录'
        assert ws.cell(1, 1).value == '报工时间'

    def test_with_records(self, tmp_path, monkeypatch):
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        records = [{
            'report_time': '2026-06-24 10:00:00',
            'workorder_no': 'WO-001',
            'order_no': 'ORD-001',
            'process_name': '织网',
            'operator_name': '王五',
            'batch_no': 'B001',
            'report_qty': 50,
            'unit': '米',
            'qualified_qty': 49,
            'loss_qty': 1,
            'equipment': '织机1号',
            'remark': '正常',
        }]
        out = build_workreport_export(records, 'wr.xlsx')
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 2).value == 'WO-001'
        assert ws.cell(2, 5).value == '王五'


class TestBuildOperatorExport:
    """build_operator_export 测试"""

    def test_empty(self, tmp_path, monkeypatch):
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        out = build_operator_export([], {}, 'empty_op.xlsx')
        assert os.path.exists(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == '操作员列表(企业架构)'
        assert ws.cell(1, 1).value == '微信UserID'

    def test_with_data(self, tmp_path, monkeypatch):
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        operators = [
            {
                'operator_id': 'wx_001',
                'name': '张三',
                'role': '员工',
                'department': '生产部',
                'phone': '13800138000',
                'enabled': True,
                'updated_at': '2026-06-24 10:00:00',
            },
            {
                'id': 'wx_002',  # 兼容 id 字段
                'name': '李四',
                'enabled': False,
            },
        ]
        stats = {
            'wx_001': {
                'total_hours': 168,
                'total_qty': 1000,
                'total_qualified': 950,
                'qualify_rate': 95.0,
            },
        }
        out = build_operator_export(operators, stats, 'op.xlsx')
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 1).value == 'wx_001'
        assert ws.cell(2, 2).value == '张三'
        assert ws.cell(2, 6).value == '在职'  # enabled=True
        assert ws.cell(3, 6).value == '停用'  # enabled=False
        assert ws.cell(2, 8).value == 168  # total_hours

    def test_default_role(self, tmp_path, monkeypatch):
        """无 role 字段时默认为'员工'"""
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        out = build_operator_export([{'operator_id': 'op1', 'name': 'A'}], {}, 'op_def.xlsx')
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 3).value == '员工'


# ── 模板生成测试 ──

class TestGenerateTemplate:
    """generate_template 测试"""

    def test_generates_template(self, tmp_path, monkeypatch):
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        out = generate_template()
        assert os.path.exists(out)
        assert '订单导入模板' in os.path.basename(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == '订单导入模板'
        # 表头 18 列
        assert ws.cell(1, 1).value == '订单号'
        # 示例数据在第 2 行
        assert ws.cell(2, 1).value == 'ORD-2026-0001'


class TestGenerateOperatorTemplate:
    """generate_operator_template 测试"""

    def test_generates(self, tmp_path, monkeypatch):
        import utils.excel_service as svc
        monkeypatch.setattr(svc, 'EXPORT_DIR', str(tmp_path))
        out = generate_operator_template()
        assert os.path.exists(out)
        assert '操作员导入模板' in os.path.basename(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == '操作员导入模板'
        assert ws.cell(1, 1).value == '微信UserID(必填)'


# ── 导入解析测试 ──

class TestParseOrderImport:
    """parse_order_import 测试"""

    def _make_excel(self, headers, rows):
        """辅助：创建 BytesIO Excel 文件"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append(r)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_empty_file(self):
        """无数据行"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['客户名称', '产品类型', '数量'])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = parse_order_import(buf)
        assert result['valid'] == []
        assert len(result['errors']) >= 1
        assert '文件为空' in result['errors'][0]['msg'] or '数据行' in result['errors'][0]['msg']

    def test_unrecognized_header(self):
        buf = self._make_excel(['未知列A', '未知列B'], [['x', 'y']])
        result = parse_order_import(buf)
        assert result['valid'] == []
        assert '未识别' in result['errors'][0]['msg']

    def test_valid_row(self):
        buf = self._make_excel(
            ['客户名称', '产品类型', '数量'],
            [['客户A', '网带', 100]]
        )
        result = parse_order_import(buf)
        assert len(result['valid']) == 1
        assert result['valid'][0]['customer_name'] == '客户A'
        assert result['valid'][0]['quantity'] == 100.0
        assert result['invalid'] == []

    def test_missing_required_field(self):
        """必填字段缺失"""
        buf = self._make_excel(
            ['客户名称', '产品类型', '数量'],
            [['', '网带', 100]]  # 客户名为空
        )
        result = parse_order_import(buf)
        assert len(result['invalid']) == 1
        assert 'customer_name' in result['invalid'][0]['missing']
        assert len(result['errors']) >= 1
        assert '缺少必填字段' in result['errors'][0]['msg']

    def test_empty_rows_skipped(self):
        """空行被跳过"""
        buf = self._make_excel(
            ['客户名称', '产品类型', '数量'],
            [
                ['客户A', '网带', 100],
                [None, None, None],  # 全空
                ['客户B', '网带', 50],
            ]
        )
        result = parse_order_import(buf)
        assert len(result['valid']) == 2
        assert result['valid'][0]['customer_name'] == '客户A'
        assert result['valid'][1]['customer_name'] == '客户B'

    def test_quantity_parsed_as_number(self):
        """quantity 解析为数字"""
        buf = self._make_excel(
            ['客户名称', '产品类型', '数量'],
            [['客户A', '网带', '50.5']]
        )
        result = parse_order_import(buf)
        assert result['valid'][0]['quantity'] == 50.5

    def test_delivery_date_parsed(self):
        """delivery_date 解析为标准日期"""
        buf = self._make_excel(
            ['客户名称', '产品类型', '数量', '交期'],
            [['客户A', '网带', 50, '2026/06/24']]
        )
        result = parse_order_import(buf)
        assert result['valid'][0]['delivery_date'] == '2026-06-24'

    def test_extra_columns_ignored(self):
        """未识别的列被忽略"""
        buf = self._make_excel(
            ['客户名称', '产品类型', '数量', '未知列'],
            [['客户A', '网带', 50, 'extra_data']]
        )
        result = parse_order_import(buf)
        assert len(result['valid']) == 1
        assert '未知列' not in result['valid'][0]


class TestParseOperatorImport:
    """parse_operator_import 测试"""

    def _make_excel(self, headers, rows):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append(r)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_empty_file(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['姓名'])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = parse_operator_import(buf)
        assert result['valid'] == []
        assert len(result['errors']) >= 1

    def test_unrecognized_header(self):
        buf = self._make_excel(['未知列'], [['x']])
        result = parse_operator_import(buf)
        assert '未识别' in result['errors'][0]['msg']

    def test_valid_row(self):
        buf = self._make_excel(
            ['姓名', '角色', '部门', '微信UserID'],
            [['张三', '员工', '生产部', 'wx_001']]
        )
        result = parse_operator_import(buf)
        assert len(result['valid']) == 1
        rec = result['valid'][0]
        assert rec['name'] == '张三'
        assert rec['role'] == '员工'
        assert rec['wechat_userid'] == 'wx_001'

    def test_default_role_and_password(self):
        """无 role/password 时填默认值"""
        buf = self._make_excel(['姓名'], [['张三']])
        result = parse_operator_import(buf)
        rec = result['valid'][0]
        assert rec['role'] == '操作员'
        assert rec['password'] == '123456'

    def test_missing_name(self):
        """必填 name 缺失"""
        buf = self._make_excel(['姓名', '部门'], [['', '生产部']])
        result = parse_operator_import(buf)
        assert len(result['invalid']) == 1
        assert 'name' in result['invalid'][0]['missing']

    def test_empty_rows_skipped(self):
        buf = self._make_excel(
            ['姓名', '部门'],
            [
                ['张三', '生产部'],
                [None, None],
                ['李四', '质检部'],
            ]
        )
        result = parse_operator_import(buf)
        assert len(result['valid']) == 2


# ── 样式常量测试 ──

class TestStyleConstants:
    """样式常量存在性测试"""

    def test_module_imports(self):
        """确保所有公开符号可导入"""
        from utils import excel_service
        # 关键函数
        assert callable(excel_service.build_order_export)
        assert callable(excel_service.build_shipment_export)
        assert callable(excel_service.build_workreport_export)
        assert callable(excel_service.build_operator_export)
        assert callable(excel_service.generate_template)
        assert callable(excel_service.generate_operator_template)
        assert callable(excel_service.parse_order_import)
        assert callable(excel_service.parse_operator_import)
