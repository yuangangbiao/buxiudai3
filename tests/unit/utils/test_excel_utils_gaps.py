# -*- coding: utf-8 -*-
"""补测 excel_utils.py 未覆盖部分: ExcelImporter, get_template_path, create_template"""
import os
import sys
import pytest
import openpyxl
from unittest.mock import patch, MagicMock, PropertyMock
from datetime import datetime


# ==============================================================
#  ExcelImporter 测试
# ==============================================================

class TestImportOrders:
    """补测 ExcelImporter.import_orders (全未覆盖)"""

    def _make_import_order_wb(self, fp, data_row):
        """Helper to create an import_orders workbook with headers + data row"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["客户名称*", "联系电话", "收货地址", "产品类型*", "材质*", "网孔(mm)",
                    "丝径(mm)", "宽度(mm)", "长度(mm)", "数量*", "单位", "单价(元)",
                    "总价(元)", "表面处理", "特殊要求", "交货日期", "状态", "备注",
                    "extra_params"])
        ws.append(data_row)
        wb.save(fp)

    def test_import_success(self, tmp_path):
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'import_orders.xlsx')

        # 代码实际索引(0-based): 0=空检查, 1=customer_name, 2=phone, 3=address,
        # 4=product_type, 5=material, 6=mesh_size, 7=wire_diameter, 8=width,
        # 9=length, 10=quantity(int!), 11=unit, 12=unit_price, 13=total_amount,
        # 14=surface_treatment, 15=special_requirements, 16=delivery_date,
        # 17=status, 18=remark, 19=extra_params — 所以数据需要至少20列
        data_row = (
            "测试客户", "13800138000", "地址1", "编织网", "304",
            5, 2, 100, 200,     # 5-8: mesh, wire_dia, width, length
            50,                  # 9: length
            100, "米",           # 10: quantity, 11: unit
            30, 1500,            # 12-13: unit_price, total_amount
            "", "", "2026-06-01", "", "", ""  # 14-19: surface... through extra_params
        )
        self._make_import_order_wb(fp, data_row)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn), \
             patch('models.database.generate_order_no', return_value='ORD-NEW-001'):
            result = ExcelImporter.import_orders(fp)

        assert result['imported'] == 1, f"imported={result['imported']}, errors={result['errors']}"
        assert result['errors'] == []
        # 验证 INSERT 调用
        mock_cursor.execute.assert_called_once()
        insert_sql = mock_cursor.execute.call_args[0][0]
        assert 'INSERT INTO orders' in insert_sql
        mock_conn.commit.assert_called_once()

    def test_import_empty_rows_skipped(self, tmp_path):
        """跳过空行"""
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'skip_empty.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["客户名称*", "联系电话", "收货地址", "产品类型*", "材质*", "网孔(mm)",
                    "丝径(mm)", "宽度(mm)", "长度(mm)", "数量*", "单位", "单价(元)",
                    "总价(元)", "表面处理", "特殊要求", "交货日期", "状态", "备注",
                    "extra_params"])
        ws.append([None] * 19)  # 空行: 19个None
        ws.append(["客户2", "13900000000", "地址2", "平网", "316",
                    3, 1, 200, 300,     # 5-8
                    30,                  # 9=length
                    30, "米", 40, 1200,   # 10-13
                    "", "", "", "", "", ""])  # 14-19
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn), \
             patch('models.database.generate_order_no', return_value='ORD-NEW-002'):
            result = ExcelImporter.import_orders(fp)

        assert result['imported'] == 1  # Only second row imported
        assert result['errors'] == []

    def test_import_row_error(self, tmp_path):
        """某行失败时收集错误"""
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'with_error.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["客户名称*", "联系电话", "收货地址", "产品类型*", "材质*", "网孔(mm)",
                    "丝径(mm)", "宽度(mm)", "长度(mm)", "数量*", "单位", "单价(元)",
                    "总价(元)", "表面处理", "特殊要求", "交货日期", "状态", "备注",
                    "extra_params"])
        ws.append(["客户1", "13800138000", "地址1", "编织网", "304",
                    5, 2, 100, 200,     # 5-8
                    50,                  # 9=length
                    100, "米", 30, 1500,  # 10-13
                    "", "", "", "", "", ""])  # 14-19
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.execute.side_effect = Exception("DB insert failed")
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn), \
             patch('models.database.generate_order_no', return_value='ORD-ERR'):
            result = ExcelImporter.import_orders(fp)

        assert result['imported'] == 0
        assert len(result['errors']) == 1
        assert '第2行' in result['errors'][0]
        assert 'DB insert failed' in result['errors'][0]


class TestImportInventory:
    """补测 ExcelImporter.import_inventory"""

    def test_import_new(self, tmp_path):
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'import_inv.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["材料名称*", "材料类型*", "规格", "当前库存", "单位",
                    "单价(元)", "仓库", "预警线", "备注"])
        ws.append(["不锈钢丝", "原材料", "φ2mm", 1000, "kg", 20, "主仓库", 100, "常用材料"])
        ws.append(["铜线", "原材料", "φ1mm", 500, "kg", 50, "主仓库", 50, ""])
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = None  # no existing record
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelImporter.import_inventory(fp)

        assert result['imported'] == 2
        assert result['errors'] == []
        # 两行都不存在 => INSERT twice
        insert_count = sum(1 for call in mock_cursor.execute.call_args_list
                           if 'INSERT INTO inventory' in call[0][0])
        assert insert_count == 2

    def test_import_update_existing(self, tmp_path):
        """已存在记录时更新"""
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'inv_update.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["材料名称*", "材料类型*", "规格", "当前库存", "单位",
                    "单价(元)", "仓库", "预警线", "备注"])
        ws.append(["不锈钢丝", "原材料", "φ2mm", 2000, "kg", 25, "主仓库", 100, "补充库存"])
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        # 第一行 fetchone() 返回已存在记录
        mock_cursor.fetchone.return_value = {"id": 1}
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelImporter.import_inventory(fp)

        assert result['imported'] == 1
        assert result['errors'] == []
        # 应为 UPDATE 而非 INSERT
        update_calls = [c for c in mock_cursor.execute.call_args_list
                        if 'UPDATE inventory' in c[0][0]]
        assert len(update_calls) == 1

    def test_import_inventory_error(self, tmp_path):
        """异常时收集错误"""
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'inv_err.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["材料名称*", "材料类型*", "规格", "当前库存", "单位",
                    "单价(元)", "仓库", "预警线", "备注"])
        ws.append(["不锈钢丝", "原材料", "φ2mm", 1000, "kg", 20, "主仓库", 100, ""])
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.execute.side_effect = Exception("库存导入失败")
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelImporter.import_inventory(fp)

        assert result['imported'] == 0
        assert len(result['errors']) == 1


class TestImportBom:
    """补测 ExcelImporter.import_bom"""

    def test_import_success(self, tmp_path):
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'import_bom.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["产品类型*", "材质*", "用钢量(kg/米)", "用钢单位", "损耗率(%)",
                    "包装材料", "表面处理", "生产工艺", "计量单位", "备注"])
        ws.append(["编织网带", "不锈钢304", 1.5, "kg/米", 5, "纸箱", "抛光", "编织→检验", "米", ""])
        ws.append(["平网", "不锈钢316", 2.0, "kg/米", 3, "木托", "酸洗", "编织→酸洗→检验", "米", ""])
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.execute.return_value = None  # INSERT ok for first attempt
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelImporter.import_bom(fp)

        assert result['imported'] == 2
        assert result['errors'] == []

    def test_import_skip_empty_row(self, tmp_path):
        """跳过空行: product_type 或 material 为空"""
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'bom_skip.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["产品类型*", "材质*", "用钢量(kg/米)", "用钢单位", "损耗率(%)",
                    "包装材料", "表面处理", "生产工艺", "计量单位", "备注"])
        ws.append([None, None, None, None, None, None, None, None, None, None])  # skipped
        ws.append(["编织网带", "不锈钢304", 1.5, "kg/米", 5, "纸箱", "抛光", "编织→检验", "米", ""])
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.execute.return_value = None
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelImporter.import_bom(fp)

        assert result['imported'] == 1

    def test_import_insert_fallback_to_update(self, tmp_path):
        """INSERT 失败时回退到 UPDATE (duplicate)"""
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'bom_dup.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["产品类型*", "材质*", "用钢量(kg/米)", "用钢单位", "损耗率(%)",
                    "包装材料", "表面处理", "生产工艺", "计量单位", "备注"])
        ws.append(["编织网带", "不锈钢304", 2.0, "kg/米", 5, "纸箱+木托", "抛光", "编织→检验→包装", "米", "更新"])
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        # 第一次 INSERT 抛异常 (duplicate)
        mock_cursor.execute.side_effect = [Exception("Duplicate"), None]
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelImporter.import_bom(fp)

        assert result['imported'] == 1
        # 应该有 UPDATEE call
        update_calls = [c for c in mock_cursor.execute.call_args_list
                        if 'UPDATE bom_list' in str(c)]
        assert len(update_calls) >= 1

    def test_import_bom_row_error(self, tmp_path):
        """整行处理抛异常时收集错误"""
        from utils.excel_utils import ExcelImporter
        fp = os.path.join(str(tmp_path), 'bom_err.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["产品类型*", "材质*", "用钢量(kg/米)", "用钢单位", "损耗率(%)",
                    "包装材料", "表面处理", "生产工艺", "计量单位", "备注"])
        ws.append(["编织网带", "不锈钢304", 1.5, "kg/米", 5, "纸箱", "抛光", "编织→检验", "米", ""])
        wb.save(fp)

        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.execute.side_effect = Exception("BOM导入失败")
        mock_conn.cursor.return_value = mock_cursor

        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelImporter.import_bom(fp)

        assert result['imported'] == 0
        assert len(result['errors']) == 1


# ==============================================================
#  get_template_path / create_template 测试
# ==============================================================

class TestGetTemplatePath:
    """补测 get_template_path"""

    def test_returns_path_with_templates_dir(self):
        from utils.excel_utils import get_template_path
        result = get_template_path("orders_template.xlsx")
        assert result.endswith("orders_template.xlsx")
        assert "templates" in result

    def test_creates_templates_dir(self, tmp_path):
        from utils.excel_utils import get_template_path, BASE_DIR
        # Temporarily override BASE_DIR
        import utils.excel_utils as eu
        orig_base = eu.BASE_DIR
        eu.BASE_DIR = str(tmp_path)
        try:
            result = get_template_path("test_template.xlsx")
            assert os.path.exists(os.path.join(str(tmp_path), "templates"))
            assert result == os.path.join(str(tmp_path), "templates", "test_template.xlsx")
        finally:
            eu.BASE_DIR = orig_base


class TestCreateTemplate:
    """补测 create_template"""

    def test_create_orders_template(self, tmp_path):
        from utils.excel_utils import create_template, BASE_DIR
        import utils.excel_utils as eu
        orig_base = eu.BASE_DIR
        eu.BASE_DIR = str(tmp_path)
        try:
            file_path = create_template("orders")
            assert os.path.exists(file_path)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            assert ws.title == "订单导入模板"
            # 有表头行 + 示例行
            assert ws.max_row >= 2
            assert ws.cell(1, 1).value == "客户名称*"
        finally:
            eu.BASE_DIR = orig_base

    def test_create_inventory_template(self, tmp_path):
        from utils.excel_utils import create_template, BASE_DIR
        import utils.excel_utils as eu
        orig_base = eu.BASE_DIR
        eu.BASE_DIR = str(tmp_path)
        try:
            file_path = create_template("inventory")
            assert os.path.exists(file_path)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            assert ws.title == "库存导入模板"
            assert ws.cell(1, 1).value == "材料名称*"
            assert ws.cell(2, 1).value == "不锈钢丝"
        finally:
            eu.BASE_DIR = orig_base

    def test_create_bom_template(self, tmp_path):
        from utils.excel_utils import create_template, BASE_DIR
        import utils.excel_utils as eu
        orig_base = eu.BASE_DIR
        eu.BASE_DIR = str(tmp_path)
        try:
            file_path = create_template("bom")
            assert os.path.exists(file_path)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            assert ws.title == "BOM导入模板"
            assert ws.cell(1, 1).value == "产品类型*"
            assert ws.cell(2, 1).value == "编织网带"
        finally:
            eu.BASE_DIR = orig_base

    def test_create_unknown_template(self, tmp_path):
        """未知模板名称，走不到任何 if 分支但仍应生成空文件"""
        from utils.excel_utils import create_template, BASE_DIR
        import utils.excel_utils as eu
        orig_base = eu.BASE_DIR
        eu.BASE_DIR = str(tmp_path)
        try:
            # get_template_path 在 makedirs 时会创建目录
            file_path = create_template("unknown")
            assert os.path.exists(file_path)
            wb = openpyxl.load_workbook(file_path)
            # 没有 if 命中，只有默认 Workbook + _auto_width + save
            assert wb.active is not None
        finally:
            eu.BASE_DIR = orig_base


# ==============================================================
#  ExcelExporter 边界场景
# ==============================================================

class TestExportMaterialPrepExtra:
    """补测 export_material_prep 未覆盖的边界"""

    def test_export_empty_prep(self, tmp_path):
        from utils.excel_utils import ExcelExporter
        fp = os.path.join(str(tmp_path), 'empty_prep.xlsx')
        mock_conn = MagicMock()
        mock_conn.cursor.return_value.fetchall.return_value = []
        with patch('models.database.get_connection', return_value=mock_conn):
            assert ExcelExporter.export_material_prep(fp) is True
        # 验证文件生成且至少包含表头
        wb = openpyxl.load_workbook(fp)
        ws = wb.active
        assert ws.title == "备料清单"
        assert ws.cell(1, 1).value == "订单号"


class TestExportOrdersExtra:
    """补测 export_orders 的 extra_params JSON 路径"""

    def test_extra_params_valid_json(self, tmp_path):
        from utils.excel_utils import ExcelExporter
        fp = os.path.join(str(tmp_path), 'ep_json.xlsx')
        row = {"order_no": "ORD-JSON", "customer_name": "A", "customer_phone": "",
               "customer_address": "", "product_type": "X", "material": "304",
               "mesh_size": 12, "wire_diameter": 2.0, "width": 100, "length": 50,
               "quantity": 100, "unit": "米", "unit_price": 15.0, "total_amount": 1500.0,
               "surface_treatment": "", "special_requirements": "",
               "delivery_date": "", "status": "pending", "remark": "",
               "extra_params": '{"key1":"val1","key2":"val2"}',
               "created_at": "2026-01-01 12:00:00", "updated_at": "2026-01-02 12:00:00"}
        mock_conn = MagicMock()
        mock_conn.cursor.return_value.fetchall.return_value = [row]
        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelExporter.export_orders(fp)
        assert result is True
        wb = openpyxl.load_workbook(fp)
        ws = wb.active
        # 第 2 行第 20 列（extra_params 列）应为 "key1=val1; key2=val2"
        assert ws.cell(2, 20).value == "key1=val1; key2=val2"

    def test_extra_params_invalid_json(self, tmp_path):
        """extra_params 是非 JSON 字符串"""
        from utils.excel_utils import ExcelExporter
        fp = os.path.join(str(tmp_path), 'ep_nojson.xlsx')
        row = {"order_no": "ORD-NOJSON", "customer_name": "A", "customer_phone": "",
               "customer_address": "", "product_type": "X", "material": "304",
               "mesh_size": 12, "wire_diameter": 2.0, "width": 100, "length": 50,
               "quantity": 100, "unit": "米", "unit_price": 15.0, "total_amount": 1500.0,
               "surface_treatment": "", "special_requirements": "",
               "delivery_date": "", "status": "pending", "remark": "",
               "extra_params": "raw_string_not_json",
               "created_at": None, "updated_at": None}
        mock_conn = MagicMock()
        mock_conn.cursor.return_value.fetchall.return_value = [row]
        with patch('models.database.get_connection', return_value=mock_conn):
            result = ExcelExporter.export_orders(fp)
        assert result is True
        wb = openpyxl.load_workbook(fp)
        ws = wb.active
        assert ws.cell(2, 20).value == "raw_string_not_json"


class TestExcelExporterClass:
    """补测类常量"""

    def test_data_font(self):
        from utils.excel_utils import ExcelExporter
        assert ExcelExporter.DATA_FONT.name == "微软雅黑"
        assert ExcelExporter.DATA_FONT.size == 10

    def test_border(self):
        from utils.excel_utils import ExcelExporter
        assert ExcelExporter.BORDER.left.style == 'thin'
        assert ExcelExporter.BORDER.right.style == 'thin'
