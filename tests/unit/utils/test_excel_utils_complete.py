# -*- coding: utf-8 -*-
"""
utils/excel_utils.py 基础单元测试

覆盖模块:
- ExcelExporter
"""
import os
import sys
import pytest
from unittest.mock import Mock, MagicMock, patch
from openpyxl import Workbook

class TestExcelExporterExists:
    """ExcelExporter 存在性测试"""

    def test_excel_exporter_class_exists(self):
        """测试ExcelExporter类存在"""
        from utils.excel_utils import ExcelExporter
        assert ExcelExporter is not None

    def test_excel_exporter_has_export_orders_method(self):
        """测试export_orders方法存在"""
        from utils.excel_utils import ExcelExporter
        assert hasattr(ExcelExporter, 'export_orders')
        assert callable(ExcelExporter.export_orders)

    def test_excel_exporter_has_set_header_method(self):
        """测试_set_header方法存在"""
        from utils.excel_utils import ExcelExporter
        assert hasattr(ExcelExporter, '_set_header')

    def test_excel_exporter_has_set_data_row_method(self):
        """测试_set_data_row方法存在"""
        from utils.excel_utils import ExcelExporter
        assert hasattr(ExcelExporter, '_set_data_row')

    def test_excel_exporter_has_auto_width_method(self):
        """测试_auto_width方法存在"""
        from utils.excel_utils import ExcelExporter
        assert hasattr(ExcelExporter, '_auto_width')


class TestExcelExporterStyles:
    """ExcelExporter 样式测试"""

    def test_excel_exporter_has_header_fill(self):
        """测试HEADER_FILL样式存在"""
        from utils.excel_utils import ExcelExporter
        assert hasattr(ExcelExporter, 'HEADER_FILL')

    def test_excel_exporter_has_header_font(self):
        """测试HEADER_FONT样式存在"""
        from utils.excel_utils import ExcelExporter
        assert hasattr(ExcelExporter, 'HEADER_FONT')

    def test_excel_exporter_has_data_font(self):
        """测试DATA_FONT样式存在"""
        from utils.excel_utils import ExcelExporter
        assert hasattr(ExcelExporter, 'DATA_FONT')

    def test_excel_exporter_has_border(self):
        """测试BORDER样式存在"""
        from utils.excel_utils import ExcelExporter
        assert hasattr(ExcelExporter, 'BORDER')


class TestExcelExporterSetHeader:
    """ExcelExporter._set_header 测试"""

    def test_set_header_with_valid_headers(self):
        """测试设置有效表头"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        headers = ["订单号", "客户名称", "数量"]
        ExcelExporter._set_header(ws, headers)

        assert ws.cell(row=1, column=1).value == "订单号"
        assert ws.cell(row=1, column=2).value == "客户名称"
        assert ws.cell(row=1, column=3).value == "数量"

    def test_set_header_sets_fill(self):
        """测试表头设置填充色"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        headers = ["测试"]
        ExcelExporter._set_header(ws, headers)

        cell = ws.cell(row=1, column=1)
        assert cell.fill is not None

    def test_set_header_sets_font(self):
        """测试表头设置字体"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        headers = ["测试"]
        ExcelExporter._set_header(ws, headers)

        cell = ws.cell(row=1, column=1)
        assert cell.font is not None

    def test_set_header_sets_alignment(self):
        """测试表头设置对齐"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        headers = ["测试"]
        ExcelExporter._set_header(ws, headers)

        cell = ws.cell(row=1, column=1)
        assert cell.alignment is not None

    def test_set_header_sets_border(self):
        """测试表头设置边框"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        headers = ["测试"]
        ExcelExporter._set_header(ws, headers)

        cell = ws.cell(row=1, column=1)
        assert cell.border is not None


class TestExcelExporterSetDataRow:
    """ExcelExporter._set_data_row 测试"""

    def test_set_data_row_with_valid_values(self):
        """测试设置有效数据行"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        values = ["ORD001", "张三", 100]
        ExcelExporter._set_data_row(ws, 2, values)

        assert ws.cell(row=2, column=1).value == "ORD001"
        assert ws.cell(row=2, column=2).value == "张三"
        assert ws.cell(row=2, column=3).value == 100

    def test_set_data_row_sets_font(self):
        """测试数据行设置字体"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        values = ["test"]
        ExcelExporter._set_data_row(ws, 2, values)

        cell = ws.cell(row=2, column=1)
        assert cell.font is not None

    def test_set_data_row_sets_alignment(self):
        """测试数据行设置对齐"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        values = ["test"]
        ExcelExporter._set_data_row(ws, 2, values)

        cell = ws.cell(row=2, column=1)
        assert cell.alignment is not None

    def test_set_data_row_sets_border(self):
        """测试数据行设置边框"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        values = ["test"]
        ExcelExporter._set_data_row(ws, 2, values)

        cell = ws.cell(row=2, column=1)
        assert cell.border is not None

    def test_set_data_row_with_empty_values(self):
        """测试设置空值"""
        from utils.excel_utils import ExcelExporter

        wb = Workbook()
        ws = wb.active

        values = ["", None, ""]
        ExcelExporter._set_data_row(ws, 2, values)

        assert ws.cell(row=2, column=1).value == ""
        assert ws.cell(row=2, column=2).value is None


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])
