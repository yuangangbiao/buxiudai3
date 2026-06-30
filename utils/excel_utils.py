# -*- coding: utf-8 -*-
"""
Excel导入导出工具
支持：订单、库存、BOM、备料的Excel导入导出
"""
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 打包路径兼容（pathlib 规范化中文路径）
if getattr(sys, 'frozen', False):
    BASE_DIR = str(__import__('pathlib').Path(sys._MEIPASS))
else:
    BASE_DIR = str(__import__('pathlib').Path(os.path.abspath(__file__)).parent.parent)


class ExcelExporter:
    """Excel导出器"""

    # 样式定义
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    DATA_FONT = Font(name="微软雅黑", size=10)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def _set_header(ws, headers: list):
        """设置表头"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ExcelExporter.BORDER

    @staticmethod
    def _set_data_row(ws, row_num: int, values: list):
        """设置数据行"""
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = ExcelExporter.DATA_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ExcelExporter.BORDER

    @staticmethod
    def _auto_width(ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def export_orders(file_path: str) -> bool:
        """导出订单到Excel"""
        import json
        from models.database import get_connection
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT order_no, customer_name, customer_phone, customer_address,
                   product_type, material, mesh_size, wire_diameter, width, length,
                   quantity, unit, unit_price, total_amount, surface_treatment,
                   special_requirements, delivery_date, status, remark, extra_params,
                   created_at, updated_at
            FROM orders ORDER BY created_at DESC
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "订单列表"

        headers = ["订单号", "客户名称", "联系电话", "收货地址", "产品类型", "材质",
                  "网孔(mm)", "丝径(mm)", "宽度(mm)", "长度(mm)", "数量",
                  "单位", "单价(元)", "总价(元)", "表面处理", "特殊要求",
                  "交货日期", "状态", "备注", "扩展参数", "创建时间", "更新时间"]

        ExcelExporter._set_header(ws, headers)

        for row_num, row in enumerate(rows, 2):
            # 解析 extra_params JSON
            ep_raw = row["extra_params"] or ""
            try:
                ep = json.loads(ep_raw)
                ep_str = "; ".join(f"{k}={v}" for k, v in ep.items()) if ep else ""
            except (json.JSONDecodeError, TypeError):
                ep_str = ep_raw
            values = [
                row["order_no"], row["customer_name"], row["customer_phone"],
                row["customer_address"], row["product_type"], row["material"],
                row["mesh_size"], row["wire_diameter"], row["width"], row["length"],
                row["quantity"], row["unit"], row["unit_price"], row["total_amount"],
                row["surface_treatment"], row["special_requirements"],
                row["delivery_date"], row["status"], row["remark"],
                ep_str,
                row["created_at"][:19] if row["created_at"] else "",
                row["updated_at"][:19] if row["updated_at"] else ""
            ]
            ExcelExporter._set_data_row(ws, row_num, values)

        ExcelExporter._auto_width(ws)
        wb.save(file_path)
        return True

    @staticmethod
    def export_inventory(file_path: str) -> bool:
        """导出库存到Excel"""
        from models.database import get_connection
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT material_name, material_type, specification, quantity,
                   unit, unit_price, warehouse, warning_qty, remark, updated_at
            FROM inventory ORDER BY material_type, material_name
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "库存列表"

        headers = ["材料名称", "材料类型", "规格", "当前库存", "单位", "单价(元)",
                  "仓库", "预警线", "备注", "更新时间"]

        ExcelExporter._set_header(ws, headers)

        for row_num, row in enumerate(rows, 2):
            values = [
                row["material_name"], row["material_type"], row["specification"],
                row["quantity"], row["unit"], row["unit_price"],
                row["warehouse"], row["warning_qty"], row["remark"],
                row["updated_at"][:19] if row["updated_at"] else ""
            ]
            ExcelExporter._set_data_row(ws, row_num, values)

        ExcelExporter._auto_width(ws)
        wb.save(file_path)
        return True

    @staticmethod
    def export_bom(file_path: str) -> bool:
        """导出BOM到Excel"""
        from models.database import get_connection
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT product_type, material, steel_weight, steel_unit, waste_rate,
                   packaging_materials, surface_treatment, production_process,
                   unit, remark, created_at
            FROM bom_list ORDER BY product_type, material
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM清单"

        headers = ["产品类型", "材质", "用钢量(kg/米)", "用钢单位", "损耗率(%)",
                  "包装材料", "表面处理", "生产工艺", "计量单位", "备注", "创建时间"]

        ExcelExporter._set_header(ws, headers)

        for row_num, row in enumerate(rows, 2):
            values = [
                row["product_type"], row["material"], row["steel_weight"],
                row["steel_unit"], row["waste_rate"], row["packaging_materials"],
                row["surface_treatment"], row["production_process"],
                row["unit"], row["remark"],
                row["created_at"][:19] if row["created_at"] else ""
            ]
            ExcelExporter._set_data_row(ws, row_num, values)

        ExcelExporter._auto_width(ws)
        wb.save(file_path)
        return True

    @staticmethod
    def export_material_prep(file_path: str) -> bool:
        """导出备料清单到Excel"""
        from models.database import get_connection
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT o.order_no, o.customer_name, om.material_name, om.material_type,
                   om.required_qty, om.prepared_qty, om.unit, om.prep_status,
                   om.warehouse, om.remark
            FROM order_materials om
            JOIN orders o ON om.order_id = o.id
            ORDER BY o.order_no, om.material_name
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "备料清单"

        headers = ["订单号", "客户", "材料名称", "材料类型", "需求数量",
                  "已备数量", "单位", "备料状态", "仓库", "备注"]

        ExcelExporter._set_header(ws, headers)

        for row_num, row in enumerate(rows, 2):
            values = [
                row["order_no"], row["customer_name"], row["material_name"],
                row["material_type"], row["required_qty"], row["prepared_qty"],
                row["unit"], row["prep_status"], row["warehouse"], row["remark"]
            ]
            ExcelExporter._set_data_row(ws, row_num, values)

        ExcelExporter._auto_width(ws)
        wb.save(file_path)
        return True


class ExcelImporter:
    """Excel导入器"""

    # [BUG 9 修复 2026-06-16] 进程内导入去重缓存，防止同一 Excel 60s 内重复导入
    _IMPORT_HISTORY = {}
    _IMPORT_COOLDOWN_SECONDS = 60

    @classmethod
    def _check_recent_duplicate(cls, file_path: str) -> int:
        """检查文件指纹是否在冷却期内 (返回 0=无重复, 1=命中重复)

        指纹算法: 文件路径 + size + mtime + md5(前 1024 字节)
        """
        import hashlib
        import time as _time
        if not os.path.exists(file_path):
            return 0
        try:
            size = os.path.getsize(file_path)
            mtime = os.path.getmtime(file_path)
            with open(file_path, 'rb') as _f:
                head_hash = hashlib.md5(_f.read(1024)).hexdigest()
        except Exception:
            return 0
        key = f'{file_path}|{size}|{mtime}|{head_hash}'
        now = _time.time()
        # 清理过期项
        cls._IMPORT_HISTORY = {
            k: v for k, v in cls._IMPORT_HISTORY.items()
            if now - v < cls._IMPORT_COOLDOWN_SECONDS
        }
        if key in cls._IMPORT_HISTORY:
            return 1
        cls._IMPORT_HISTORY[key] = now
        return 0

    @staticmethod
    def import_orders(file_path: str) -> dict:
        """从Excel导入订单"""
        # [BUG 9 修复 2026-06-16] 同一 Excel 60s 内重复导入直接拒绝，防止 230 条测试数据再次灌入
        dup = ExcelImporter._check_recent_duplicate(file_path)
        if dup:
            return {
                'imported': 0,
                'errors': [f'文件 {os.path.basename(file_path)} 在 {ExcelImporter._IMPORT_COOLDOWN_SECONDS}s 内已导入过，请稍后再试或换文件'],
                'duplicate': True,
            }

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 获取表头
        headers = [cell.value for cell in ws[1]]

        imported = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:  # 跳过空行
                continue
            try:
                from models.database import get_connection, generate_order_no
                conn = get_connection()
                cursor = conn.cursor()

                data = {
                    "order_no": generate_order_no(),
                    "customer_name": str(row[1] or ""),
                    "customer_phone": str(row[2] or ""),
                    "customer_address": str(row[3] or ""),
                    "product_type": str(row[4] or ""),
                    "material": str(row[5] or ""),
                    "mesh_size": float(row[6]) if row[6] else None,
                    "wire_diameter": float(row[7]) if row[7] else None,
                    "width": float(row[8]) if row[8] else None,
                    "length": float(row[9]) if row[9] else None,
                    "quantity": int(row[10]) if row[10] else 1,
                    "unit": str(row[11] or "米"),
                    "unit_price": float(row[12]) if row[12] else 0,
                    "total_amount": float(row[13]) if row[13] else 0,
                    "surface_treatment": str(row[14] or ""),
                    "special_requirements": str(row[15] or ""),
                    "delivery_date": str(row[16])[:10] if row[16] else None,
                    "status": str(row[17]) if row[17] else "待确认",
                    "remark": str(row[18]) if row[18] else "",
                    "extra_params": str(row[19]) if row[19] else "",
                }

                cursor.execute("""
                    INSERT INTO orders (order_no, customer_name, customer_phone, customer_address,
                        product_type, material, mesh_size, wire_diameter, width, length,
                        quantity, unit, unit_price, total_amount, surface_treatment,
                        special_requirements, delivery_date, status, remark, extra_params)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (data["order_no"], data["customer_name"], data["customer_phone"], data["customer_address"],
                      data["product_type"], data["material"], data["mesh_size"], data["wire_diameter"],
                      data["width"], data["length"], data["quantity"], data["unit"], data["unit_price"],
                      data["total_amount"], data["surface_treatment"], data["special_requirements"],
                      data["delivery_date"], data["status"], data["remark"], data["extra_params"]))
                conn.commit()
                cursor.close()
                conn.close()
                imported += 1

            except Exception as e:
                errors.append(f"第{row_num}行: {str(e)}")

        return {"imported": imported, "errors": errors}

    @staticmethod
    def import_inventory(file_path: str) -> dict:
        """从Excel导入库存"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        imported = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:  # 跳过空行
                continue
            try:
                from models.database import get_connection
                conn = get_connection()
                cursor = conn.cursor()

                material_name = str(row[0] or "")
                material_type = str(row[1] or "原材料")
                specification = str(row[2] or "")
                quantity = float(row[3]) if row[3] else 0
                unit = str(row[4] or "kg")
                unit_price = float(row[5]) if row[5] else 0
                warehouse = str(row[6] or "主仓库")
                warning_qty = float(row[7]) if row[7] else 50
                remark = str(row[8]) if row[8] else ""
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # 检查是否已存在
                cursor.execute(
                    "SELECT id FROM inventory WHERE material_name = %s AND warehouse = %s",
                    (material_name, warehouse)
                )
                exists = cursor.fetchone()

                if exists:
                    # 更新
                    cursor.execute("""
                        UPDATE inventory SET material_type=%s, specification=%s, quantity=%s,
                            unit=%s, unit_price=%s, warning_qty=%s, remark=%s,
                            updated_at=%s
                        WHERE id = %s
                    """, (material_type, specification, quantity,
                          unit, unit_price, warning_qty,
                          remark, now, exists["id"]))
                else:
                    cursor.execute("""
                        INSERT INTO inventory (material_name, material_type, specification,
                            quantity, unit, unit_price, warehouse, warning_qty, remark)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (material_name, material_type, specification,
                          quantity, unit, unit_price, warehouse, warning_qty, remark))

                conn.commit()
                cursor.close()
                conn.close()
                imported += 1

            except Exception as e:
                errors.append(f"第{row_num}行: {str(e)}")

        return {"imported": imported, "errors": errors}

    @staticmethod
    def import_bom(file_path: str) -> dict:
        """从Excel导入BOM"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        imported = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0] or not row[1]:  # 跳过空行
                continue
            try:
                from models.database import get_connection
                conn = get_connection()
                cursor = conn.cursor()
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                product_type = str(row[0] or "")
                material = str(row[1] or "")
                steel_weight = float(row[2]) if row[2] else 0
                steel_unit = str(row[3] or "kg/米")
                waste_rate = float(row[4]) if row[4] else 5
                packaging_materials = str(row[5]) if row[5] else ""
                surface_treatment = str(row[6]) if row[6] else ""
                production_process = str(row[7]) if row[7] else ""
                unit = str(row[8] or "米")
                remark = str(row[9]) if row[9] else ""

                # 尝试插入，失败则更新
                try:
                    cursor.execute("""
                        INSERT INTO bom_list (product_type, material, steel_weight, steel_unit,
                            waste_rate, packaging_materials, surface_treatment, production_process,
                            unit, remark)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (product_type, material, steel_weight, steel_unit,
                          waste_rate, packaging_materials, surface_treatment,
                          production_process, unit, remark))
                except Exception:
                    # 更新已存在的记录
                    cursor.execute("""
                        UPDATE bom_list SET steel_weight=%s, steel_unit=%s, waste_rate=%s,
                            packaging_materials=%s, surface_treatment=%s, production_process=%s,
                            unit=%s, remark=%s, updated_at=%s
                        WHERE product_type=%s AND material=%s
                    """, (steel_weight, steel_unit, waste_rate,
                          packaging_materials, surface_treatment,
                          production_process, unit, remark,
                          now, product_type, material))

                conn.commit()
                cursor.close()
                conn.close()
                imported += 1

            except Exception as e:
                errors.append(f"第{row_num}行: {str(e)}")

        return {"imported": imported, "errors": errors}


def get_template_path(template_name: str) -> str:
    """获取模板文件路径"""
    templates_dir = os.path.join(BASE_DIR, "templates")
    os.makedirs(templates_dir, exist_ok=True)
    return os.path.join(templates_dir, template_name)


def create_template(template_name: str) -> str:
    """创建Excel模板"""
    wb = openpyxl.Workbook()
    ws = wb.active

    if template_name == "orders":
        ws.title = "订单导入模板"
        headers = ["客户名称*", "联系电话", "收货地址", "产品类型*", "材质*", "网孔(mm)",
                  "丝径(mm)", "宽度(mm)", "长度(mm)", "数量*", "单位", "单价(元)",
                  "总价(元)", "表面处理", "特殊要求", "交货日期", "状态", "备注"]
        ws.append(headers)

        # 添加示例行
        example = ["示例客户", "13800138000", "示例地址", "编织网带", "不锈钢304",
                  5.0, 1.0, 1000, 2000, 100, "米", 50, 5000, "抛光", "无", "2026-05-01", "待确认", "测试订单"]
        ws.append(example)

    elif template_name == "inventory":
        ws.title = "库存导入模板"
        headers = ["材料名称*", "材料类型*", "规格", "当前库存", "单位",
                  "单价(元)", "仓库", "预警线", "备注"]
        ws.append(headers)

        example = ["不锈钢丝", "原材料", "φ2mm", 1000, "kg", 20, "主仓库", 100, "常用材料"]
        ws.append(example)

    elif template_name == "bom":
        ws.title = "BOM导入模板"
        headers = ["产品类型*", "材质*", "用钢量(kg/米)", "用钢单位", "损耗率(%)",
                  "包装材料", "表面处理", "生产工艺", "计量单位", "备注"]
        ws.append(headers)

        example = ["编织网带", "不锈钢304", 1.5, "kg/米", 5, "纸箱+木托", "抛光", "编织→检验→包装", "米", ""]
        ws.append(example)

    ExcelExporter._auto_width(ws)

    file_path = get_template_path(f"{template_name}_template.xlsx")
    wb.save(file_path)
    return file_path
