# -*- coding: utf-8 -*-
"""
Excel 导入导出服务（订单 + 发货单 + 报工记录）
"""
import os
import json
import tempfile
import logging
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT_DIR = os.path.join(PROJECT_ROOT, 'exports')
os.makedirs(EXPORT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', start_color='4472C4', end_color='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
NORMAL_FONT = Font(size=10)
DATE_FILL = PatternFill('solid', start_color='E8F4FD', end_color='E8F4FD')
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')


def _format_val(v: Any) -> Any:
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else str(v)
    return v


def _auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 4, min_width), max_width
        )


def _write_header_row(ws, row_num, headers: List[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _write_data_row(ws, row_num, values: List[Any], aligns: Optional[List[str]]=None):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=_format_val(v))
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if aligns and col <= len(aligns):
            a = aligns[col - 1]
            cell.alignment = CENTER_ALIGN if a == 'center' else (RIGHT_ALIGN if a == 'right' else LEFT_ALIGN)
        else:
            cell.alignment = LEFT_ALIGN


def build_order_export(orders: List[Dict], filename: str) -> str:
    headers = [
        '订单号', '产品类型', '客户名称', '客户分组',
        '材质', '网孔尺寸', '丝径(mm)', '宽度(mm)', '长度(m)',
        '数量', '单位', '单价(元)', '总价(元)',
        '表面处理', '特殊要求', '交期', '状态',
        '备注', '客户电话', '客户地址',
        '创建时间',
    ]
    aligns = [
        'left', 'left', 'left', 'left',
        'center', 'left', 'right', 'right', 'right',
        'right', 'center', 'right', 'right',
        'left', 'left', 'center', 'center',
        'left', 'left', 'left',
        'center',
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = '订单列表'
    ws.freeze_panes = 'A2'
    _write_header_row(ws, 1, headers)

    for i, o in enumerate(orders, 2):
        extra = o.get('extra_params', {}) or {}
        if isinstance(extra, str):
            try:
                extra = json.loads(extra)
            except Exception:
                extra = {}
        _write_data_row(ws, i, [
            o.get('order_no', ''),
            o.get('product_type', '') or o.get('product_name', ''),
            o.get('customer_name', ''),
            o.get('customer_group', ''),
            o.get('material', ''),
            o.get('mesh_size', ''),
            o.get('wire_diameter', ''),
            o.get('width', ''),
            o.get('length', ''),
            o.get('quantity', 0),
            o.get('unit', '米'),
            o.get('unit_price', ''),
            o.get('total_amount', o.get('total', '')),
            o.get('surface_treatment', ''),
            o.get('special_requirements', '') or o.get('product_remark', ''),
            (o.get('delivery_date') or '')[:10] if o.get('delivery_date') else '',
            o.get('order_status', '') or o.get('status', ''),
            o.get('remark', ''),
            o.get('customer_phone', ''),
            o.get('customer_address', ''),
            (o.get('created_at') or '')[:19] if o.get('created_at') else '',
        ], aligns)

    _auto_width(ws)
    safe_name = ''.join(c for c in filename if c not in ('/', '\\', ':', '*', '?', '"', '<', '>', '|'))
    out_path = os.path.join(EXPORT_DIR, safe_name)
    wb.save(out_path)
    logger.info(f'[Excel导出] 订单 {len(orders)} 条 → {out_path}')
    return out_path


def generate_template() -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = '订单导入模板'
    headers = [
        '订单号', '产品类型', '客户名称', '客户分组', '客户电话', '客户地址',
        '材质', '网孔尺寸', '丝径(mm)', '宽度(mm)', '长度(m)',
        '数量', '单位', '单价(元)', '表面处理', '特殊要求', '交期(YYYY-MM-DD)', '备注',
    ]
    _write_header_row(ws, 1, headers)
    example = [
        'ORD-2026-0001', '普通网带', '示例客户', '华东', '13800138000', '示例地址',
        '304', '10mm×10mm', '2.0', '1000', '50',
        '100', '米', '50.0', '抛光', '无', datetime.now().strftime('%Y-%m-%d'), '这是示例数据',
    ]
    for col, v in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.font = Font(size=10, color='808080')
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'A2'
    _auto_width(ws)
    out_path = os.path.join(EXPORT_DIR, '订单导入模板.xlsx')
    wb.save(out_path)
    logger.info(f'[Excel模板] 生成 → {out_path}')
    return out_path


REQUIRED_FIELDS = ['customer_name', 'product_type', 'quantity']
FIELD_ALIASES = {
    '客户名称': 'customer_name', '客户': 'customer_name',
    '产品类型': 'product_type', '产品': 'product_type',
    '数量': 'quantity', '订单数量': 'quantity', 'qty': 'quantity',
    '订单号': 'order_no',
    '客户分组': 'customer_group', '分组': 'customer_group',
    '客户电话': 'customer_phone', '电话': 'customer_phone',
    '客户地址': 'customer_address', '地址': 'customer_address',
    '材质': 'material',
    '网孔尺寸': 'mesh_size',
    '丝径': 'wire_diameter', '丝径(mm)': 'wire_diameter',
    '宽度': 'width', '宽度(mm)': 'width',
    '长度': 'length', '长度(m)': 'length',
    '单位': 'unit',
    '单价': 'unit_price', '单价(元)': 'unit_price',
    '表面处理': 'surface_treatment',
    '特殊要求': 'special_requirements', '产品备注': 'product_remark',
    '交期': 'delivery_date', '交期(YYYY-MM-DD)': 'delivery_date',
    '备注': 'remark',
}


def _normalize_header(h: str) -> Optional[str]:
    h = h.strip()
    if h in FIELD_ALIASES:
        return FIELD_ALIASES[h]
    for alias, norm in FIELD_ALIASES.items():
        if alias in h or h in alias:
            return norm
    return None


def _parse_date(val: Any) -> Optional[str]:
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return s


def _parse_num(val: Any) -> Any:
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    try:
        return float(s)
    except Exception:
        return s


def parse_order_import(file_storage_obj) -> Dict[str, Any]:
    wb = load_workbook(file_storage_obj, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {'valid': [], 'invalid': [], 'errors': [{'row': 0, 'msg': '文件为空或无数据行'}]}
    raw_headers = rows[0]
    header_map = {}
    for ci, h in enumerate(raw_headers):
        if h:
            norm = _normalize_header(str(h))
            if norm:
                header_map[ci] = norm
    if not header_map:
        return {'valid': [], 'invalid': [], 'errors': [{'row': 1, 'msg': '未识别到有效表头，请使用标准模板'}]}

    valid, invalid, errors = [], [], []
    for ri, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        record = {}
        missing = []
        raw = {}
        for ci, norm in header_map.items():
            raw[norm] = row[ci]
            if norm in REQUIRED_FIELDS:
                v = row[ci]
                if v is None or str(v).strip() == '':
                    missing.append(norm)
                else:
                    record[norm] = _parse_num(v) if norm in ('quantity', 'unit_price', 'width', 'length', 'wire_diameter') else _parse_date(v) if norm in ('delivery_date',) else str(v).strip()
            else:
                v = row[ci]
                if v is not None and str(v).strip() != '':
                    record[norm] = _parse_num(v) if norm in ('quantity', 'unit_price', 'width', 'length', 'wire_diameter') else _parse_date(v) if norm in ('delivery_date',) else str(v).strip()
        if missing:
            invalid.append({'row': ri, 'data': raw, 'missing': missing})
            errors.append({'row': ri, 'msg': f'缺少必填字段: {", ".join(missing)}'})
        else:
            valid.append(record)
    return {'valid': valid, 'invalid': invalid, 'errors': errors}


def build_shipment_export(records: List[Dict], filename: str) -> str:
    headers = ['发货单号', '订单号', '客户名称', '产品类型', '发货数量', '单位',
               '发货日期', '收货人', '收货地址', '物流公司', '运单号', '发货人', '状态', '创建时间']
    aligns = ['left', 'left', 'left', 'left', 'right', 'center', 'center',
              'left', 'left', 'left', 'left', 'left', 'center', 'center']
    wb = Workbook()
    ws = wb.active
    ws.title = '发货记录'
    ws.freeze_panes = 'A2'
    _write_header_row(ws, 1, headers)
    for i, r in enumerate(records, 2):
        _write_data_row(ws, i, [
            r.get('shipment_no', ''), r.get('order_no', ''),
            r.get('customer_name', ''), r.get('product_type', ''),
            r.get('quantity', 0), r.get('unit', '米'),
            (r.get('ship_date') or '')[:10] if r.get('ship_date') else '',
            r.get('receiver', ''), r.get('receiver_address', ''),
            r.get('logistics_company', ''), r.get('tracking_no', ''),
            r.get('shipper', ''),
            r.get('status', ''),
            (r.get('created_at') or '')[:19] if r.get('created_at') else '',
        ], aligns)
    _auto_width(ws)
    out_path = os.path.join(EXPORT_DIR, filename)
    wb.save(out_path)
    logger.info(f'[Excel导出] 发货记录 {len(records)} 条 → {out_path}')
    return out_path


def build_workreport_export(records: List[Dict], filename: str) -> str:
    headers = ['报工时间', '工单号', '订单号', '工序', '操作员', '批次号', '报工数量', '单位',
               '合格数量', '损耗数量', '设备', '备注']
    aligns = ['center', 'left', 'left', 'left', 'left', 'left', 'right', 'center',
              'right', 'right', 'left', 'left']
    wb = Workbook()
    ws = wb.active
    ws.title = '报工记录'
    ws.freeze_panes = 'A2'
    _write_header_row(ws, 1, headers)
    for i, r in enumerate(records, 2):
        _write_data_row(ws, i, [
            (r.get('report_time') or '')[:19] if r.get('report_time') else '',
            r.get('workorder_no', ''), r.get('order_no', ''),
            r.get('process_name', ''), r.get('operator_name', ''),
            r.get('batch_no', ''),
            r.get('report_qty', 0), r.get('unit', '米'),
            r.get('qualified_qty', 0), r.get('loss_qty', 0),
            r.get('equipment', ''), r.get('remark', ''),
        ], aligns)
    _auto_width(ws)
    out_path = os.path.join(EXPORT_DIR, filename)
    wb.save(out_path)
    logger.info(f'[Excel导出] 报工记录 {len(records)} 条 → {out_path}')
    return out_path


def _format_dt(v) -> str:
    if v is None or v == '':
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return str(v)[:19]


def build_operator_export(operators: List[Dict], stats_map: Dict[str, Dict], filename: str) -> str:
    headers = ['微信UserID', '姓名', '角色', '部门', '电话', '状态', '更新时间',
               '本月工时(小时)', '本月报工数', '本月合格数', '合格率(%)']
    aligns = ['left', 'left', 'center', 'left', 'left', 'center', 'center',
              'right', 'right', 'right', 'right']
    wb = Workbook()
    ws = wb.active
    ws.title = '操作员列表(企业架构)'
    ws.freeze_panes = 'A2'
    _write_header_row(ws, 1, headers)
    for i, o in enumerate(operators, 2):
        op_id = o.get('operator_id') or o.get('id', '')
        stats = stats_map.get(op_id, {})
        enabled = o.get('enabled', True)
        status = '在职' if enabled else '停用'
        _write_data_row(ws, i, [
            op_id,
            o.get('name', ''),
            o.get('role', '员工'),
            o.get('department', ''),
            o.get('phone', ''),
            status,
            _format_dt(o.get('updated_at')),
            stats.get('total_hours', 0),
            stats.get('total_qty', 0),
            stats.get('total_qualified', 0),
            stats.get('qualify_rate', 0),
        ], aligns)
    _auto_width(ws)
    safe_name = ''.join(c for c in filename if c not in ('/', '\\', ':', '*', '?', '"', '<', '>', '|'))
    out_path = os.path.join(EXPORT_DIR, safe_name)
    wb.save(out_path)
    logger.info(f'[Excel导出] 操作员 {len(operators)} 条 → {out_path}')
    return out_path


def generate_operator_template() -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = '操作员导入模板'
    headers = ['微信UserID(必填)', '姓名(必填)', '角色(员工/经理/管理员)', '部门(从企业架构选)', '电话', '密码(选填,默认123456)']
    _write_header_row(ws, 1, headers)
    example = ['weixin_001', '示例员工', '员工', '生产部', '13800138000', '123456']
    for col, v in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.font = Font(size=10, color='808080')
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'A2'
    _auto_width(ws)
    out_path = os.path.join(EXPORT_DIR, '操作员导入模板.xlsx')
    wb.save(out_path)
    logger.info(f'[Excel模板] 操作员 → {out_path}')
    return out_path


OPERATOR_REQUIRED_FIELDS = ['name']
OPERATOR_FIELD_ALIASES = {
    '姓名': 'name', '操作员姓名': 'name',
    '操作员ID': 'operator_id', '工号': 'operator_id',
    '角色': 'role', '权限': 'role',
    '部门': 'department',
    '微信UserID': 'wechat_userid', '微信号': 'wechat_userid',
    '密码': 'password', '初始密码': 'password',
}


def _normalize_operator_header(h: str) -> Optional[str]:
    h = h.strip()
    if h in OPERATOR_FIELD_ALIASES:
        return OPERATOR_FIELD_ALIASES[h]
    for alias, norm in OPERATOR_FIELD_ALIASES.items():
        if alias in h or h in alias:
            return norm
    return None


def parse_operator_import(file_storage_obj) -> Dict[str, Any]:
    wb = load_workbook(file_storage_obj, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {'valid': [], 'invalid': [], 'errors': [{'row': 0, 'msg': '文件为空或无数据行'}]}
    header_map = {}
    for ci, h in enumerate(rows[0]):
        if h:
            norm = _normalize_operator_header(str(h))
            if norm:
                header_map[ci] = norm
    if not header_map:
        return {'valid': [], 'invalid': [], 'errors': [{'row': 1, 'msg': '未识别到有效表头，请使用标准模板'}]}

    valid, invalid, errors = [], [], []
    for ri, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        record = {}
        missing = []
        for ci, norm in header_map.items():
            v = row[ci]
            if norm in OPERATOR_REQUIRED_FIELDS:
                if v is None or str(v).strip() == '':
                    missing.append(norm)
                else:
                    record[norm] = str(v).strip()
            else:
                if v is not None and str(v).strip() != '':
                    record[norm] = str(v).strip()
        if missing:
            invalid.append({'row': ri, 'missing': missing})
            errors.append({'row': ri, 'msg': f'缺少必填字段: {", ".join(missing)}'})
        else:
            if not record.get('role'):
                record['role'] = '操作员'
            if not record.get('password'):
                record['password'] = '123456'
            valid.append(record)
    return {'valid': valid, 'invalid': invalid, 'errors': errors}
